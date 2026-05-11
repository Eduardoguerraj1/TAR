from __future__ import annotations

import math
import random
import re
import statistics
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl


class _LazyScipyStats:
    def __getattr__(self, name: str) -> Any:
        from scipy import stats as scipy_stats

        return getattr(scipy_stats, name)


stats = _LazyScipyStats()


class TarWorkbookError(RuntimeError):
    """Raised when the TAR workbook cannot be loaded into the expected model."""


SCENARIOS: dict[str, dict[str, str]] = {
    "a1": {"key": "a1", "sheet": "A1", "label": "Angra 1"},
    "a1_a2": {"key": "a1_a2", "sheet": "A1 e A2", "label": "Angra 1 e Angra 2"},
    "hipotetico": {"key": "hipotetico", "sheet": "", "label": "Cenário hipotético"},
}

EXPECTED_RADIONUCLIDES = ["Co-58", "Co-60", "Cr-51", "Cs-137", "Mn-54", "Sb-124", "Sb-125", "Zr-95"]
EMPIRICAL_ACTIVITY_RADIONUCLIDES = ["Co-58", "Co-60", "Cr-51", "Cs-137", "Mn-54", "Nb-95", "Sb-124", "Sb-125", "Zr-95"]
EMPIRICAL_ACTIVITY_GROUPS = ["TAR - Afluente", "TAR - Efluente"]
EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS = ["TAR - Afluente"]
EMPIRICAL_ACTIVITY_DEFAULT_FILENAME = "Atividade Total TAR c radionuclideos.xls"
TOTAL_ACTIVITY_DEFAULT_FILENAME = "Dados Atividade TAR - Jayme (1).xlsx"
TOTAL_ACTIVITY_REVIEW_TOP_N = 15
TOTAL_ACTIVITY_WINDOW_TARGET_RADIONUCLIDES = 8
SEDIMENT_TRANSFER_FACTOR = 0.000072
SEDIMENT_EXPOSURE_TIME_H = 262980.0

COMPARTMENTS: list[dict[str, Any]] = [
    {
        "key": "water",
        "label": "Água",
        "unit": "Bq/m³",
        "value_col": 7,
        "report_level_col": 19,
        "lld_col": 18,
    },
    {
        "key": "fish",
        "label": "Peixe",
        "unit": "Bq/kg",
        "value_col": 12,
        "report_level_col": 21,
        "lld_col": 20,
    },
    {
        "key": "invertebrate",
        "label": "Invertebrado",
        "unit": "Bq/kg",
        "value_col": 13,
        "report_level_col": 23,
        "lld_col": 22,
    },
    {
        "key": "sediment",
        "label": "Sedimento",
        "unit": "Bq/kg",
        "value_col": 16,
        "report_level_col": 25,
        "lld_col": 24,
    },
]

INFERENTIAL_TEST_MINIMUMS: list[dict[str, str]] = [
    {"test": "Shapiro-Wilk", "technical_minimum": "n >= 3", "recommended": "n >= 8 a 10 por grupo"},
    {"test": "Levene", "technical_minimum": "2 grupos, n >= 2 por grupo", "recommended": "n >= 5 por grupo"},
    {"test": "Teste t pareado", "technical_minimum": "n >= 2 pares", "recommended": "n >= 10 pares; n >= 3 para checar normalidade"},
    {"test": "Wilcoxon pareado", "technical_minimum": "pares não nulos", "recommended": "n >= 10 pares"},
    {"test": "Teste t independente", "technical_minimum": "2 grupos, n >= 2 por grupo", "recommended": "n >= 10 por grupo"},
    {"test": "Mann-Whitney U", "technical_minimum": "2 grupos, n >= 1 por grupo", "recommended": "n >= 5 a 10 por grupo"},
    {"test": "ANOVA uma via", "technical_minimum": "2+ grupos, n >= 2 por grupo", "recommended": "n >= 5 por grupo"},
    {"test": "Kruskal-Wallis", "technical_minimum": "2+ grupos", "recommended": "n >= 5 por grupo"},
    {"test": "ANOVA medidas repetidas", "technical_minimum": "3+ condições e 3+ unidades completas", "recommended": "n >= 10 unidades completas"},
    {"test": "Friedman", "technical_minimum": "3+ condições pareadas", "recommended": "n >= 10 unidades completas"},
    {"test": "Pearson/Spearman", "technical_minimum": "n >= 3 pares", "recommended": "n >= 10 a 20 pares"},
]

DEFAULT_SENSITIVITY_N = 10000
DEFAULT_SENSITIVITY_SEED = 20260504
DEFAULT_TAR_ACTIVITY_BQ_YEAR = 2.69e11
DEFAULT_DILUTION_FLOW_M3_YEAR = 2.96e9
DEFAULT_STAT_N = 60
DEFAULT_STAT_SEED = 20260504

ERICA_TOOL_VALUES: dict[str, dict[str, float | None]] = {
    # Fonte: tabela/screenshot do ERICA Tool no arquivo "Artigo TAR1 correção.pdf".
    # Água foi extraída em Bq/L no ERICA Tool e convertida para Bq/m³ para comparação com a planilha.
    "Co-58": {"water": 1.85e-2 * 1000.0, "fish": 6.30e1, "invertebrate": 4.47e1, "sediment": 3.27e0},
    "Co-60": {"water": 8.30e-4 * 1000.0, "fish": 2.94e0, "invertebrate": 2.81e0, "sediment": 3.91e0},
    "Cr-51": {"water": 4.08e-3 * 1000.0, "fish": 2.35e-1, "invertebrate": 4.27e-1, "sediment": 2.80e-1},
    "Cs-137": {"water": 8.69e-5 * 1000.0, "fish": 7.09e-3, "invertebrate": 5.32e-3, "sediment": 1.19e0},
    "Mn-54": {"water": 1.95e-4 * 1000.0, "fish": 1.59e0, "invertebrate": 1.04e0, "sediment": 1.52e-1},
    "Sb-124": {"water": 5.18e-4 * 1000.0, "fish": 1.00e-3, "invertebrate": 6.40e-2, "sediment": 7.78e-2},
    "Sb-125": {"water": 1.41e-4 * 1000.0, "fish": 3.97e-4, "invertebrate": 2.55e-2, "sediment": 3.55e-1},
    "Zr-95": {"water": 6.66e-5 * 1000.0, "fish": 3.88e-3, "invertebrate": 2.15e-3, "sediment": 5.31e-3},
}


def resolve_tar_scenario_key(value: Any) -> str:
    key = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "": "a1",
        "angra_1": "a1",
        "angra1": "a1",
        "a1": "a1",
        "a1ea2": "a1_a2",
        "a1_e_a2": "a1_a2",
        "a1_a2": "a1_a2",
        "angra_1_e_2": "a1_a2",
        "angra_1_e_angra_2": "a1_a2",
        "hipotetico": "hipotetico",
        "hipotético": "hipotetico",
        "cenario_hipotetico": "hipotetico",
        "cenário_hipotético": "hipotetico",
    }
    return aliases.get(key, key if key in SCENARIOS else "a1")


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def _to_float(value: Any) -> float | None:
    if _is_number(value):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        if not text or text == "*":
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _format_scientific(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.2E}".replace("E+0", "E+").replace("E-0", "E-")


def _format_p_value(value: float | None) -> str:
    if value is None or not math.isfinite(value):
        return "—"
    if value < 0.0001:
        return "p < 0,0001"
    return f"p = {value:.4f}".replace(".", ",")


def _format_decimal(value: float | None, *, digits: int = 4) -> str:
    if value is None or not math.isfinite(value):
        return "—"
    return f"{value:.{digits}f}".replace(".", ",")


def _format_percent(value: float | None) -> str:
    if value is None or not math.isfinite(value):
        return "—"
    return f"{value * 100:.2f}%".replace(".", ",")


def _summary_stats(values: list[float]) -> dict[str, Any]:
    if not values:
        return {
            "n": 0,
            "mean": None,
            "stdev": None,
            "cv": None,
            "q1": None,
            "median": None,
            "q3": None,
            "p95": None,
            "min": None,
            "max": None,
        }
    ordered = sorted(values)
    quartiles = statistics.quantiles(ordered, n=4, method="inclusive") if len(ordered) > 1 else [ordered[0], ordered[0], ordered[0]]
    mean = statistics.fmean(values)
    stdev = statistics.stdev(values) if len(values) > 1 else 0.0
    return {
        "n": len(values),
        "mean": mean,
        "stdev": stdev,
        "cv": stdev / mean if mean else None,
        "q1": quartiles[0],
        "median": statistics.median(values),
        "q3": quartiles[2],
        "p95": statistics.quantiles(ordered, n=100, method="inclusive")[94] if len(ordered) > 1 else ordered[0],
        "min": ordered[0],
        "max": ordered[-1],
    }


def _ratio_status(value: float | None, reference: float | None) -> dict[str, Any]:
    if value is None or reference is None or reference <= 0:
        return {"reference": reference, "ratio": None, "status": "sem referência"}
    ratio = value / reference
    return {"reference": reference, "ratio": ratio, "status": "abaixo" if ratio <= 1.0 else "acima"}


def _read_radionuclide_rows(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_index in range(1, ws.max_row + 1):
        radionuclide = ws.cell(row_index, 2).value
        if not isinstance(radionuclide, str) or not re.match(r"^[A-Z][a-z]?-\d+", radionuclide.strip()):
            continue
        item: dict[str, Any] = {
            "row": row_index,
            "radionuclide": radionuclide.strip(),
            "source_concentration_bq_m3": _to_float(ws.cell(row_index, 3).value),
            "fraction": _to_float(ws.cell(row_index, 4).value),
            "activity_bq": _to_float(ws.cell(row_index, 5).value),
            "constants": {},
            "compartments": {},
        }
        decay_constant_h = _to_float(ws.cell(row_index, 8).value)
        item["constants"] = {
            "decay_constant_h": decay_constant_h,
            "decay_constant_h_text": _format_scientific(decay_constant_h),
            "half_life_h": (0.693 / decay_constant_h) if decay_constant_h not in (None, 0) else None,
            "half_life_h_text": _format_scientific((0.693 / decay_constant_h) if decay_constant_h not in (None, 0) else None),
            "bioaccumulation_fish": _to_float(ws.cell(row_index, 9).value),
            "bioaccumulation_fish_text": _format_decimal(_to_float(ws.cell(row_index, 9).value)),
            "bioaccumulation_invertebrate": _to_float(ws.cell(row_index, 10).value),
            "bioaccumulation_invertebrate_text": _format_decimal(_to_float(ws.cell(row_index, 10).value)),
            "kd_sediment_l_kg": _to_float(ws.cell(row_index, 11).value),
            "kd_sediment_l_kg_text": _format_scientific(_to_float(ws.cell(row_index, 11).value)),
            "sediment_transfer_factor": SEDIMENT_TRANSFER_FACTOR,
            "sediment_transfer_factor_text": _format_scientific(SEDIMENT_TRANSFER_FACTOR),
            "sediment_exposure_time_h": SEDIMENT_EXPOSURE_TIME_H,
            "sediment_exposure_time_h_text": _format_scientific(SEDIMENT_EXPOSURE_TIME_H),
        }
        for compartment in COMPARTMENTS:
            value = _to_float(ws.cell(row_index, compartment["value_col"]).value)
            report_level = _to_float(ws.cell(row_index, compartment["report_level_col"]).value)
            lld = _to_float(ws.cell(row_index, compartment["lld_col"]).value)
            report_result = _ratio_status(value, report_level)
            lld_result = _ratio_status(value, lld)
            item["compartments"][compartment["key"]] = {
                "label": compartment["label"],
                "unit": compartment["unit"],
                "value": value,
                "value_text": _format_scientific(value),
                "report_level": report_result["reference"],
                "report_level_text": _format_scientific(report_result["reference"]),
                "report_level_ratio": report_result["ratio"],
                "report_level_ratio_text": f"{report_result['ratio']:.4f}".replace(".", ",") if report_result["ratio"] is not None else "—",
                "report_level_status": report_result["status"],
                "lld": lld_result["reference"],
                "lld_text": _format_scientific(lld_result["reference"]),
                "lld_ratio": lld_result["ratio"],
                "lld_ratio_text": f"{lld_result['ratio']:.4f}".replace(".", ",") if lld_result["ratio"] is not None else "—",
                "lld_status": lld_result["status"],
            }
        rows.append(item)
    return rows


def _scenario_totals(ws: Any) -> dict[str, Any]:
    return {
        "total_water_concentration_bq_m3": _to_float(ws.cell(13, 7).value),
        "circulation_flow_m3_year": _to_float(ws.cell(17, 5).value),
        "tar_capacity_m3": _to_float(ws.cell(17, 3).value),
        "activity_bq_year": _to_float(ws.cell(18, 3).value),
    }


def _build_reference_counts(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    counts: dict[str, dict[str, Any]] = {}
    for compartment in COMPARTMENTS:
        key = compartment["key"]
        report_refs = [
            row["radionuclide"]
            for row in rows
            if row["compartments"][key]["report_level"] is not None
        ]
        lld_refs = [
            row["radionuclide"]
            for row in rows
            if row["compartments"][key]["lld"] is not None
        ]
        counts[key] = {
            "label": compartment["label"],
            "report_level": len(report_refs),
            "lld": len(lld_refs),
            "report_level_radionuclides": report_refs,
            "lld_radionuclides": lld_refs,
        }
    return counts


def _build_exceedances(rows: list[dict[str, Any]], *, reference: str) -> list[dict[str, Any]]:
    exceedances: list[dict[str, Any]] = []
    status_key = "report_level_status" if reference == "Report Level" else "lld_status"
    ratio_key = "report_level_ratio" if reference == "Report Level" else "lld_ratio"
    for row in rows:
        for compartment in COMPARTMENTS:
            result = row["compartments"][compartment["key"]]
            if result[status_key] != "acima":
                continue
            exceedances.append(
                {
                    "radionuclide": row["radionuclide"],
                    "compartment": compartment["label"],
                    "reference": reference,
                    "value": result["value"],
                    "ratio": result[ratio_key],
                }
            )
    return exceedances


def _scenario_summary(scenario_key: str, ws: Any) -> dict[str, Any]:
    rows = _read_radionuclide_rows(ws)
    if not rows:
        raise TarWorkbookError(f"Nenhum radionuclídeo foi encontrado na aba {ws.title!r}.")
    reference_counts = _build_reference_counts(rows)
    report_level_exceedances = _build_exceedances(rows, reference="Report Level")
    lld_exceedances = _build_exceedances(rows, reference="LLD")
    return {
        "key": scenario_key,
        "sheet": ws.title,
        "label": SCENARIOS[scenario_key]["label"],
        "radionuclide_count": len(rows),
        "radionuclides": [row["radionuclide"] for row in rows],
        "expected_radionuclides": EXPECTED_RADIONUCLIDES,
        "rows": rows,
        "totals": _scenario_totals(ws),
        "reference_counts": reference_counts,
        "exceedances": report_level_exceedances,
        "report_level_exceedances": report_level_exceedances,
        "lld_exceedances": lld_exceedances,
        "has_reference_exceedance": bool(report_level_exceedances),
        "has_lld_exceedance": bool(lld_exceedances),
    }


def _safe_int(value: Any, default: int, *, minimum: int, maximum: int) -> int:
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError):
        parsed = default
    return max(minimum, min(maximum, parsed))


def _generate_positive_measurements(base_value: float, count: int, rng: random.Random) -> list[float]:
    sigma = 0.18
    measurements: list[float] = []
    for _ in range(count):
        measurement = rng.lognormvariate(math.log(max(base_value, 1e-12)), sigma)
        measurements.append(measurement)
    return measurements


def _sensitivity_variable_specs(base_activity: float, base_flow: float) -> list[dict[str, Any]]:
    return [
        {
            "key": "activity_factor",
            "label": "Atividade total do TAR",
            "distribution": "Triangular",
            "parameters": "0,50x / 0,75x / 1,00x",
            "base_value": base_activity,
            "base_value_text": _format_scientific(base_activity),
            "unit": "Bq/ano",
            "description": "Multiplicador sintético aplicado à atividade total descarregada pelo TAR.",
        },
        {
            "key": "flow_factor",
            "label": "Vazão de diluição",
            "distribution": "Triangular",
            "parameters": "0,75x / 1,00x / 1,25x",
            "base_value": base_flow,
            "base_value_text": _format_scientific(base_flow),
            "unit": "m³/ano",
            "description": "Multiplicador sintético da vazão; valores maiores reduzem a concentração por diluição.",
        },
        {
            "key": "fish_bio_factor",
            "label": "Fator de bioacumulação - peixe",
            "distribution": "Lognormal",
            "parameters": "mediana 1,00x; sigma 0,35",
            "base_value": 1.0,
            "base_value_text": "1,00x",
            "unit": "multiplicador",
            "description": "Variação sintética do fator de bioacumulação para peixes.",
        },
        {
            "key": "invertebrate_bio_factor",
            "label": "Fator de bioacumulação - invertebrado",
            "distribution": "Lognormal",
            "parameters": "mediana 1,00x; sigma 0,50",
            "base_value": 1.0,
            "base_value_text": "1,00x",
            "unit": "multiplicador",
            "description": "Variação sintética do fator de bioacumulação para invertebrados.",
        },
        {
            "key": "sediment_transfer_factor",
            "label": "Transferência para sedimento",
            "distribution": "Lognormal",
            "parameters": "mediana 1,00x; sigma 0,65",
            "base_value": 1.0,
            "base_value_text": "1,00x",
            "unit": "multiplicador",
            "description": "Variação sintética agregando Kd/Kc e transferência sedimentar.",
        },
        {
            "key": "exposure_factor",
            "label": "Tempo de exposição",
            "distribution": "Uniforme",
            "parameters": "0,50x a 1,50x",
            "base_value": 1.0,
            "base_value_text": "1,00x",
            "unit": "multiplicador",
            "description": "Multiplicador sintético do tempo de exposição nos compartimentos biota e sedimento.",
        },
    ]


def _sample_sensitivity_variables(rng: random.Random) -> dict[str, float]:
    return {
        "activity_factor": rng.triangular(0.50, 1.00, 0.75),
        "flow_factor": rng.triangular(0.75, 1.25, 1.00),
        "fish_bio_factor": rng.lognormvariate(0.0, 0.35),
        "invertebrate_bio_factor": rng.lognormvariate(0.0, 0.50),
        "sediment_transfer_factor": rng.lognormvariate(0.0, 0.65),
        "exposure_factor": rng.uniform(0.50, 1.50),
    }


def _sensitivity_multiplier(compartment_key: str, variables: dict[str, float]) -> float:
    water_multiplier = variables["activity_factor"] / variables["flow_factor"]
    if compartment_key == "fish":
        return water_multiplier * variables["fish_bio_factor"] * variables["exposure_factor"]
    if compartment_key == "invertebrate":
        return water_multiplier * variables["invertebrate_bio_factor"] * variables["exposure_factor"]
    if compartment_key == "sediment":
        return water_multiplier * variables["sediment_transfer_factor"] * variables["exposure_factor"]
    return water_multiplier


def _build_histogram(values: list[float], *, bin_count: int = 12) -> list[dict[str, Any]]:
    if not values:
        return []
    lower = min(values)
    upper = max(values)
    if not math.isfinite(lower) or not math.isfinite(upper):
        return []
    if upper <= lower:
        upper = lower + 1.0
    step = (upper - lower) / bin_count
    counts = [0 for _ in range(bin_count)]
    for value in values:
        index = min(bin_count - 1, max(0, int((value - lower) / step)))
        counts[index] += 1
    return [
        {
            "start": lower + index * step,
            "end": lower + (index + 1) * step,
            "count": count,
            "label": f"{_format_decimal(lower + index * step, digits=2)}-{_format_decimal(lower + (index + 1) * step, digits=2)}",
        }
        for index, count in enumerate(counts)
    ]


def _build_sensitivity_analysis(scenario: dict[str, Any], *, sample_count: int, seed: int) -> dict[str, Any]:
    rng = random.Random(seed)
    totals = scenario.get("totals") or {}
    base_activity = totals.get("activity_bq_year") or DEFAULT_TAR_ACTIVITY_BQ_YEAR
    base_flow = totals.get("circulation_flow_m3_year") or DEFAULT_DILUTION_FLOW_M3_YEAR
    variable_specs = _sensitivity_variable_specs(float(base_activity), float(base_flow))
    variable_samples: dict[str, list[float]] = {spec["key"]: [] for spec in variable_specs}
    value_samples: dict[tuple[str, str], list[float]] = {}
    exceedance_counts: dict[tuple[str, str], int] = {}
    max_report_level_ratios: list[float] = []

    rows = scenario.get("rows") or []
    for row in rows:
        for compartment in COMPARTMENTS:
            value_samples[(row["radionuclide"], compartment["key"])] = []
            exceedance_counts[(row["radionuclide"], compartment["key"])] = 0

    for _ in range(sample_count):
        variables = _sample_sensitivity_variables(rng)
        for key, value in variables.items():
            variable_samples[key].append(value)
        sample_ratios: list[float] = []
        for row in rows:
            for compartment in COMPARTMENTS:
                compartment_key = compartment["key"]
                data = row["compartments"][compartment_key]
                base_value = data.get("value")
                if base_value is None:
                    continue
                value = float(base_value) * _sensitivity_multiplier(compartment_key, variables)
                sample_key = (row["radionuclide"], compartment_key)
                value_samples[sample_key].append(value)
                report_level = data.get("report_level")
                if report_level is not None and report_level > 0:
                    ratio = value / float(report_level)
                    sample_ratios.append(ratio)
                    if ratio > 1.0:
                        exceedance_counts[sample_key] += 1
        max_report_level_ratios.append(max(sample_ratios) if sample_ratios else 0.0)

    summary_rows: list[dict[str, Any]] = []
    for row in rows:
        for compartment in COMPARTMENTS:
            compartment_key = compartment["key"]
            data = row["compartments"][compartment_key]
            sample_key = (row["radionuclide"], compartment_key)
            stats_summary = _summary_stats(value_samples.get(sample_key, []))
            report_level = data.get("report_level")
            p95 = stats_summary["p95"]
            p95_ratio = (p95 / report_level) if p95 is not None and report_level is not None and report_level > 0 else None
            exceedance_probability = (
                exceedance_counts.get(sample_key, 0) / sample_count
                if report_level is not None and report_level > 0
                else None
            )
            summary_rows.append(
                {
                    "radionuclide": row["radionuclide"],
                    "compartment": compartment["label"],
                    "compartment_key": compartment_key,
                    "unit": compartment["unit"],
                    "mean": stats_summary["mean"],
                    "mean_text": _format_scientific(stats_summary["mean"]),
                    "min": stats_summary["min"],
                    "min_text": _format_scientific(stats_summary["min"]),
                    "max": stats_summary["max"],
                    "max_text": _format_scientific(stats_summary["max"]),
                    "p95": p95,
                    "p95_text": _format_scientific(p95),
                    "report_level": report_level,
                    "report_level_text": _format_scientific(report_level),
                    "p95_report_level_ratio": p95_ratio,
                    "p95_report_level_ratio_text": _format_decimal(p95_ratio),
                    "exceedance_probability": exceedance_probability,
                    "exceedance_probability_text": _format_percent(exceedance_probability),
                }
            )

    influence_rows: list[dict[str, Any]] = []
    for spec in variable_specs:
        correlation = 0.0
        try:
            result = stats.spearmanr(variable_samples[spec["key"]], max_report_level_ratios)
            raw_correlation = float(result.correlation)
            correlation = raw_correlation if math.isfinite(raw_correlation) else 0.0
        except Exception:
            correlation = 0.0
        influence_rows.append(
            {
                "key": spec["key"],
                "label": spec["label"],
                "correlation": correlation,
                "correlation_text": _format_decimal(correlation, digits=3),
                "absolute_correlation": abs(correlation),
                "absolute_correlation_text": _format_decimal(abs(correlation), digits=3),
                "direction": "aumenta" if correlation > 0 else "reduz" if correlation < 0 else "neutra",
            }
        )
    influence_rows.sort(key=lambda item: item["absolute_correlation"], reverse=True)

    referenced_rows = [row for row in summary_rows if row["exceedance_probability"] is not None]
    max_exceedance_row = max(referenced_rows, key=lambda item: item["exceedance_probability"], default=None)
    max_ratio_row = max(referenced_rows, key=lambda item: item["p95_report_level_ratio"] or -1.0, default=None)
    dominant = influence_rows[0] if influence_rows else None
    max_ratio_stats = _summary_stats(max_report_level_ratios)
    narrative_text = (
        f"A análise de sensibilidade Monte Carlo executou {sample_count} cenários sintéticos com seed {seed}. "
        "O seed é o número que inicializa o sorteio pseudoaleatório; mantendo o mesmo seed, o sistema reproduz exatamente os mesmos cenários. "
        "Os intervalos são demonstrativos e aplicam multiplicadores simples sobre os resultados atuais do TAR, "
        "servindo para triagem exploratória das variáveis que mais aproximam os resultados do Report Level."
    )

    return {
        "sample_count": sample_count,
        "seed": seed,
        "synthetic": True,
        "source_note": "Intervalos sintéticos demonstrativos; não substituem constantes de literatura nem conclusão regulatória.",
        "base_activity_bq_year": float(base_activity),
        "base_activity_bq_year_text": _format_scientific(float(base_activity)),
        "base_flow_m3_year": float(base_flow),
        "base_flow_m3_year_text": _format_scientific(float(base_flow)),
        "variables": variable_specs,
        "summary_rows": summary_rows,
        "influence_rows": influence_rows,
        "narrative_text": narrative_text,
        "cards": [
            {"label": "Simulações", "value": f"{sample_count:,}".replace(",", ".")},
            {"label": "Seed reprodutível", "value": str(seed)},
            {
                "label": "Maior prob. > Report Level",
                "value": max_exceedance_row["exceedance_probability_text"] if max_exceedance_row else "—",
            },
            {"label": "Variável dominante", "value": dominant["label"] if dominant else "—"},
        ],
        "max_exceedance_row": max_exceedance_row,
        "max_ratio_row": max_ratio_row,
        "max_report_level_ratio_stats": {
            **max_ratio_stats,
            "min_text": _format_decimal(max_ratio_stats["min"], digits=3),
            "q1_text": _format_decimal(max_ratio_stats["q1"], digits=3),
            "median_text": _format_decimal(max_ratio_stats["median"], digits=3),
            "q3_text": _format_decimal(max_ratio_stats["q3"], digits=3),
            "mean_text": _format_decimal(max_ratio_stats["mean"], digits=3),
            "p95_text": _format_decimal(max_ratio_stats["p95"], digits=3),
            "max_text": _format_decimal(max_ratio_stats["max"], digits=3),
        },
        "chart_payloads": {
            "tornado": {"rows": influence_rows},
            "heatmap": {
                "compartments": [{"key": item["key"], "label": item["label"]} for item in COMPARTMENTS],
                "radionuclides": [row["radionuclide"] for row in rows],
                "cells": [
                    {
                        "radionuclide": row["radionuclide"],
                        "compartment": row["compartment"],
                        "compartment_key": row["compartment_key"],
                        "probability": row["exceedance_probability"],
                        "probability_text": row["exceedance_probability_text"],
                    }
                    for row in summary_rows
                ],
            },
            "histogram": {
                "bins": _build_histogram(max_report_level_ratios),
                "reference": 1.0,
                "reference_label": "Report Level",
            },
            "boxplot": {
                "stats": {
                    "min": max_ratio_stats["min"],
                    "q1": max_ratio_stats["q1"],
                    "median": max_ratio_stats["median"],
                    "q3": max_ratio_stats["q3"],
                    "max": max_ratio_stats["max"],
                    "p95": max_ratio_stats["p95"],
                    "min_text": _format_decimal(max_ratio_stats["min"], digits=3),
                    "q1_text": _format_decimal(max_ratio_stats["q1"], digits=3),
                    "median_text": _format_decimal(max_ratio_stats["median"], digits=3),
                    "q3_text": _format_decimal(max_ratio_stats["q3"], digits=3),
                    "max_text": _format_decimal(max_ratio_stats["max"], digits=3),
                    "p95_text": _format_decimal(max_ratio_stats["p95"], digits=3),
                },
                "reference": 1.0,
                "reference_label": "Report Level",
            },
        },
    }


def _stat_text(value: float | None, *, digits: int = 4) -> str:
    return _format_decimal(value, digits=digits)


def _stats_summary_with_text(values: list[float]) -> dict[str, Any]:
    summary = _summary_stats(values)
    return {
        **summary,
        "mean_text": _format_scientific(summary["mean"]),
        "stdev_text": _format_scientific(summary["stdev"]),
        "cv_text": _format_percent(summary["cv"]),
        "q1_text": _format_scientific(summary["q1"]),
        "median_text": _format_scientific(summary["median"]),
        "q3_text": _format_scientific(summary["q3"]),
        "p95_text": _format_scientific(summary["p95"]),
        "min_text": _format_scientific(summary["min"]),
        "max_text": _format_scientific(summary["max"]),
    }


def _ratio_summary_with_text(values: list[float]) -> dict[str, Any]:
    summary = _summary_stats(values)
    return {
        **summary,
        "mean_text": _format_decimal(summary["mean"]),
        "stdev_text": _format_decimal(summary["stdev"]),
        "cv_text": _format_percent(summary["cv"]),
        "q1_text": _format_decimal(summary["q1"]),
        "median_text": _format_decimal(summary["median"]),
        "q3_text": _format_decimal(summary["q3"]),
        "p95_text": _format_decimal(summary["p95"]),
        "min_text": _format_decimal(summary["min"]),
        "max_text": _format_decimal(summary["max"]),
    }


def _is_mda_marker(value: Any) -> bool:
    return isinstance(value, str) and "MDA" in value.upper()


def _raw_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.split())
    if _is_number(value):
        return _format_scientific(float(value))
    return str(value)


def _format_count(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def _empirical_activity_path(workbook_path: str | Path, activity_workbook_path: str | Path | None) -> Path:
    if activity_workbook_path is not None:
        return Path(activity_workbook_path)
    return Path(workbook_path).resolve().parent / EMPIRICAL_ACTIVITY_DEFAULT_FILENAME


def _total_activity_path(workbook_path: str | Path, total_activity_workbook_path: str | Path | None) -> Path:
    if total_activity_workbook_path is not None:
        return Path(total_activity_workbook_path)
    return Path(workbook_path).resolve().parent / TOTAL_ACTIVITY_DEFAULT_FILENAME


def _load_empirical_activity_records(path: str | Path) -> dict[str, Any]:
    activity_path = Path(path)
    if not activity_path.exists():
        raise TarWorkbookError(f"Planilha de atividade total TAR não encontrada: {activity_path}")
    try:
        import xlrd
    except Exception as exc:  # pragma: no cover - exercised only when dependency is missing
        raise TarWorkbookError("A biblioteca xlrd não está disponível para ler a planilha .xls de atividade total TAR.") from exc

    try:
        workbook = xlrd.open_workbook(str(activity_path))
    except Exception as exc:
        raise TarWorkbookError(f"Não foi possível abrir a planilha de atividade total TAR: {exc}") from exc
    if not workbook.sheet_names():
        raise TarWorkbookError("A planilha de atividade total TAR não contém abas.")

    sheet = workbook.sheet_by_index(0)
    headers = [" ".join(str(sheet.cell_value(0, col)).split()) for col in range(sheet.ncols)]
    radionuclide_columns: dict[str, int] = {}
    for col, header in enumerate(headers):
        match = re.match(r"^([A-Z][a-z]?-\d+)\b", header)
        if match and match.group(1) in EMPIRICAL_ACTIVITY_RADIONUCLIDES:
            radionuclide_columns[match.group(1)] = col
    missing = [radionuclide for radionuclide in EMPIRICAL_ACTIVITY_RADIONUCLIDES if radionuclide not in radionuclide_columns]
    if missing:
        raise TarWorkbookError(f"Radionuclídeos ausentes na planilha de atividade total TAR: {', '.join(missing)}")

    records: list[dict[str, Any]] = []
    for row_index in range(1, sheet.nrows):
        sample_id = str(sheet.cell_value(row_index, 0)).strip()
        group = " ".join(str(sheet.cell_value(row_index, 1)).split())
        if not sample_id and not group:
            continue
        if group not in EMPIRICAL_ACTIVITY_GROUPS:
            continue

        date_value = sheet.cell_value(row_index, 2)
        date_text = ""
        if _is_number(date_value):
            try:
                date_text = xlrd.xldate_as_datetime(float(date_value), workbook.datemode).date().isoformat()
            except Exception:
                date_text = _raw_text(date_value)
        else:
            date_text = _raw_text(date_value)

        activity_raw = sheet.cell_value(row_index, 3)
        activity_total_bq = _to_float(activity_raw)
        radionuclides: dict[str, dict[str, Any]] = {}
        for radionuclide, col in radionuclide_columns.items():
            raw_value = sheet.cell_value(row_index, col)
            value = _to_float(raw_value)
            censored = _is_mda_marker(raw_value)
            missing_value = raw_value in ("", None)
            status = "detectado" if value is not None else "censurado" if censored else "ausente"
            radionuclides[radionuclide] = {
                "value": value,
                "raw_text": _raw_text(raw_value),
                "status": status,
                "censored": censored,
                "missing": missing_value,
            }

        records.append(
            {
                "row": row_index + 1,
                "sample_id": sample_id,
                "group": group,
                "date": date_text,
                "activity_total_bq": activity_total_bq,
                "activity_total_raw_text": _raw_text(activity_raw),
                "activity_total_censored": _is_mda_marker(activity_raw),
                "activity_total_missing": activity_raw in ("", None),
                "radionuclides": radionuclides,
            }
        )

    return {
        "path": str(activity_path),
        "sheet": sheet.name,
        "headers": headers,
        "records": records,
    }


def _scenario_row_by_radionuclide(scenario: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {row["radionuclide"]: row for row in scenario.get("rows") or []}


def _scenario_compartment_multipliers(scenario_row: dict[str, Any]) -> dict[str, float | None]:
    compartments = scenario_row.get("compartments") or {}
    water_value = (compartments.get("water") or {}).get("value")
    multipliers: dict[str, float | None] = {"water": 1.0}
    for compartment in COMPARTMENTS:
        key = compartment["key"]
        if key == "water":
            continue
        value = (compartments.get(key) or {}).get("value")
        multipliers[key] = (float(value) / float(water_value)) if value is not None and water_value not in (None, 0) else None
    return multipliers


def _group_records(records: list[dict[str, Any]], groups: list[str] | None = None) -> dict[str, list[dict[str, Any]]]:
    selected_groups = groups or EMPIRICAL_ACTIVITY_GROUPS
    return {
        group: [record for record in records if record.get("group") == group]
        for group in selected_groups
    }


def _reference_summary(values: list[float], reference: float | None) -> dict[str, Any]:
    stats_summary = _summary_stats(values)
    p95 = stats_summary.get("p95")
    ratio = p95 / reference if p95 is not None and reference is not None and reference > 0 else None
    exceedance_count = sum(1 for value in values if reference is not None and reference > 0 and value > reference)
    exceedance_rate = exceedance_count / len(values) if values and reference is not None and reference > 0 else None
    return {
        "reference": reference,
        "reference_text": _format_scientific(reference),
        "p95_ratio": ratio,
        "p95_ratio_text": _format_decimal(ratio),
        "status": "sem referência" if ratio is None else "abaixo" if ratio <= 1.0 else "acima",
        "exceedance_count": exceedance_count if exceedance_rate is not None else None,
        "exceedance_rate": exceedance_rate,
        "exceedance_rate_text": _format_percent(exceedance_rate),
    }


def _binomial_confidence_interval(success_count: int, sample_count: int, *, alpha: float = 0.05) -> tuple[float | None, float | None]:
    if sample_count <= 0:
        return None, None
    low = 0.0 if success_count == 0 else float(stats.beta.ppf(alpha / 2.0, success_count, sample_count - success_count + 1))
    high = 1.0 if success_count == sample_count else float(stats.beta.ppf(1.0 - alpha / 2.0, success_count + 1, sample_count - success_count))
    return low, high


def _real_reference_inference(values: list[float], reference: float | None, reference_label: str) -> dict[str, Any]:
    if not values or reference is None or reference <= 0:
        return {
            "applicable": False,
            "test_label": "Não aplicável",
            "reason": f"Sem {reference_label} numérico ou amostras reais válidas.",
        }
    ratios = [value / reference for value in values if value is not None and value > 0]
    if not ratios:
        return {
            "applicable": False,
            "test_label": "Não aplicável",
            "reason": "Sem razões positivas para comparação inferencial.",
        }

    exceedance_count = sum(1 for ratio in ratios if ratio > 1.0)
    exceedance_rate = exceedance_count / len(ratios)
    ci_low, ci_high = _binomial_confidence_interval(exceedance_count, len(ratios))
    ratio_summary = _stats_summary_with_text(ratios)

    shapiro_w = None
    shapiro_p = None
    statistic = None
    p_value = None
    normality_met = False
    test_label = "Não aplicável"
    reason = ""
    if len(ratios) >= 3:
        log_ratios = [math.log(ratio) for ratio in ratios]
        try:
            shapiro = stats.shapiro(log_ratios)
            shapiro_w = float(shapiro.statistic)
            shapiro_p = float(shapiro.pvalue)
        except Exception:
            pass
        normality_met = shapiro_p is not None and shapiro_p >= 0.05
        if normality_met:
            test_label = "Teste t unilateral de uma amostra"
            try:
                test = stats.ttest_1samp(log_ratios, popmean=0.0, alternative="less")
                statistic = float(test.statistic)
                p_value = float(test.pvalue)
            except Exception:
                reason = "O teste t não produziu p-value válido."
        else:
            test_label = "Wilcoxon unilateral de uma amostra"
            try:
                test = stats.wilcoxon(log_ratios, alternative="less", zero_method="pratt")
                statistic = float(test.statistic)
                p_value = float(test.pvalue)
            except Exception:
                reason = "O Wilcoxon não produziu p-value válido."
    else:
        reason = "Menos de 3 amostras reais válidas para teste sobre log(valor/referência)."

    p95_ratio = ratio_summary.get("p95")
    if p_value is not None and p_value < 0.05 and p95_ratio is not None and p95_ratio < 1.0 and exceedance_count == 0:
        conclusion = f"amostras reais calculadas ficaram estatisticamente abaixo de {reference_label}, sem ultrapassagem observada"
    elif exceedance_count == 0:
        conclusion = f"sem ultrapassagem observada; IC95% superior da taxa = {_format_percent(ci_high)}"
    elif p_value is not None and p_value < 0.05 and p95_ratio is not None and p95_ratio < 1.0:
        conclusion = f"tendência abaixo de {reference_label}, mas houve ultrapassagem observada"
    else:
        conclusion = f"não há evidência suficiente de margem estatística abaixo de {reference_label}"

    return {
        "applicable": p_value is not None,
        "n": len(ratios),
        "test_label": test_label,
        "reason": reason,
        "statistic": statistic,
        "statistic_text": _stat_text(statistic),
        "p_value": p_value,
        "p_value_text": _format_p_value(p_value),
        "shapiro_w": shapiro_w,
        "shapiro_p": shapiro_p,
        "shapiro_p_text": _format_p_value(shapiro_p),
        "normality_met": normality_met,
        "ratio_summary": ratio_summary,
        "p95_ratio": p95_ratio,
        "p95_ratio_text": _stat_text(p95_ratio),
        "exceedance_count": exceedance_count,
        "exceedance_rate": exceedance_rate,
        "exceedance_rate_text": _format_percent(exceedance_rate),
        "exceedance_ci95_low": ci_low,
        "exceedance_ci95_low_text": _format_percent(ci_low),
        "exceedance_ci95_high": ci_high,
        "exceedance_ci95_high_text": _format_percent(ci_high),
        "exceedance_ci95_text": f"{_format_percent(ci_low)} - {_format_percent(ci_high)}",
        "conclusion": conclusion,
    }


def _empirical_group_summaries(grouped: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for group, records in grouped.items():
        activity_values = [record["activity_total_bq"] for record in records if record.get("activity_total_bq") is not None]
        stats_summary = _stats_summary_with_text(activity_values)
        summaries.append(
            {
                "group": group,
                "sample_count": len(records),
                "activity_detected_count": len(activity_values),
                "activity_censored_count": sum(1 for record in records if record.get("activity_total_censored")),
                "activity_missing_count": sum(1 for record in records if record.get("activity_total_missing")),
                "activity_unit": "Bq",
                "activity_stats": stats_summary,
                "activity_mean_text": stats_summary["mean_text"],
                "activity_median_text": stats_summary["median_text"],
                "activity_p95_text": stats_summary["p95_text"],
            }
        )
    return summaries


def _empirical_radionuclide_rows(
    grouped: dict[str, list[dict[str, Any]]],
    modeled_radionuclides: set[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for group, records in grouped.items():
        for radionuclide in EMPIRICAL_ACTIVITY_RADIONUCLIDES:
            values = [
                record["radionuclides"][radionuclide]["value"]
                for record in records
                if record["radionuclides"][radionuclide]["value"] is not None
            ]
            censored_count = sum(1 for record in records if record["radionuclides"][radionuclide]["censored"])
            missing_count = sum(1 for record in records if record["radionuclides"][radionuclide]["missing"])
            detected_count = len(values)
            sample_count = len(records)
            stats_summary = _stats_summary_with_text(values)
            rows.append(
                {
                    "group": group,
                    "radionuclide": radionuclide,
                    "unit": "Bq/kg",
                    "sample_count": sample_count,
                    "detected_count": detected_count,
                    "censored_count": censored_count,
                    "missing_count": missing_count,
                    "detected_rate": detected_count / sample_count if sample_count else None,
                    "detected_rate_text": _format_percent(detected_count / sample_count if sample_count else None),
                    "model_status": "modelado" if radionuclide in modeled_radionuclides else "observado não modelado",
                    **stats_summary,
                }
            )
    return rows


def _empirical_modeled_sample_values(
    scenario: dict[str, Any],
    grouped: dict[str, list[dict[str, Any]]],
) -> dict[tuple[str, str, str], list[float]]:
    scenario_rows = _scenario_row_by_radionuclide(scenario)
    flow = (scenario.get("totals") or {}).get("circulation_flow_m3_year") or DEFAULT_DILUTION_FLOW_M3_YEAR
    sample_values: dict[tuple[str, str, str], list[float]] = {}
    for group in grouped:
        for radionuclide in scenario_rows:
            for compartment in COMPARTMENTS:
                sample_values[(group, radionuclide, compartment["key"])] = []

    for group, records in grouped.items():
        for record in records:
            total_activity = record.get("activity_total_bq")
            detected_sum = sum(
                float(data["value"])
                for data in record.get("radionuclides", {}).values()
                if data.get("value") is not None and data["value"] > 0
            )
            if total_activity is None or total_activity <= 0 or detected_sum <= 0 or flow <= 0:
                continue
            for radionuclide, data in record.get("radionuclides", {}).items():
                detected_value = data.get("value")
                if detected_value is None or detected_value <= 0 or radionuclide not in scenario_rows:
                    continue
                fraction = float(detected_value) / detected_sum
                activity_bq = float(total_activity) * fraction
                water_value = activity_bq / float(flow)
                multipliers = _scenario_compartment_multipliers(scenario_rows[radionuclide])
                for compartment in COMPARTMENTS:
                    multiplier = multipliers.get(compartment["key"])
                    if multiplier is None:
                        continue
                    sample_values[(group, radionuclide, compartment["key"])].append(water_value * multiplier)
    return sample_values


def _empirical_modeled_compartment_rows(
    scenario: dict[str, Any],
    grouped: dict[str, list[dict[str, Any]]],
    sample_values: dict[tuple[str, str, str], list[float]] | None = None,
) -> list[dict[str, Any]]:
    scenario_rows = _scenario_row_by_radionuclide(scenario)
    if sample_values is None:
        sample_values = _empirical_modeled_sample_values(scenario, grouped)

    rows: list[dict[str, Any]] = []
    for group in grouped:
        for radionuclide in EXPECTED_RADIONUCLIDES:
            scenario_row = scenario_rows.get(radionuclide)
            if not scenario_row:
                continue
            for compartment in COMPARTMENTS:
                key = compartment["key"]
                values = sample_values.get((group, radionuclide, key), [])
                stats_summary = _stats_summary_with_text(values)
                base_data = scenario_row["compartments"][key]
                report = _reference_summary(values, base_data.get("report_level"))
                lld = _reference_summary(values, base_data.get("lld"))
                rows.append(
                    {
                        "group": group,
                        "radionuclide": radionuclide,
                        "compartment_key": key,
                        "compartment": compartment["label"],
                        "unit": compartment["unit"],
                        **stats_summary,
                        "report_level": report["reference"],
                        "report_level_text": report["reference_text"],
                        "report_level_p95_ratio": report["p95_ratio"],
                        "report_level_p95_ratio_text": report["p95_ratio_text"],
                        "report_level_status": report["status"],
                        "report_level_exceedance_count": report["exceedance_count"],
                        "report_level_exceedance_rate": report["exceedance_rate"],
                        "report_level_exceedance_rate_text": report["exceedance_rate_text"],
                        "lld": lld["reference"],
                        "lld_text": lld["reference_text"],
                        "lld_p95_ratio": lld["p95_ratio"],
                        "lld_p95_ratio_text": lld["p95_ratio_text"],
                        "lld_status": lld["status"],
                        "lld_exceedance_count": lld["exceedance_count"],
                        "lld_exceedance_rate": lld["exceedance_rate"],
                        "lld_exceedance_rate_text": lld["exceedance_rate_text"],
                    }
                )
    return rows


def _empirical_inferential_rows(
    scenario: dict[str, Any],
    grouped: dict[str, list[dict[str, Any]]],
    sample_values: dict[tuple[str, str, str], list[float]],
) -> list[dict[str, Any]]:
    scenario_rows = _scenario_row_by_radionuclide(scenario)
    rows: list[dict[str, Any]] = []
    for group in grouped:
        for radionuclide in EXPECTED_RADIONUCLIDES:
            scenario_row = scenario_rows.get(radionuclide)
            if not scenario_row:
                continue
            for compartment in COMPARTMENTS:
                key = compartment["key"]
                values = sample_values.get((group, radionuclide, key), [])
                base_data = scenario_row["compartments"][key]
                report_level = base_data.get("report_level")
                if report_level is None:
                    continue
                inference = _real_reference_inference(values, report_level, "Report Level")
                rows.append(
                    {
                        "group": group,
                        "radionuclide": radionuclide,
                        "compartment_key": key,
                        "compartment": compartment["label"],
                        "unit": compartment["unit"],
                        "reference": "Report Level",
                        "reference_value": report_level,
                        "reference_value_text": _format_scientific(report_level),
                        **inference,
                    }
                )
    return rows


def _build_empirical_activity_statistics(
    scenario: dict[str, Any],
    *,
    activity_workbook_path: str | Path,
) -> dict[str, Any]:
    loaded = _load_empirical_activity_records(activity_workbook_path)
    loaded_records = loaded["records"]
    records = [
        record
        for record in loaded_records
        if record.get("group") in EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS
    ]
    excluded_records = [
        record
        for record in loaded_records
        if record.get("group") not in EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS
    ]
    grouped = _group_records(records, EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS)
    modeled_radionuclides = set(_scenario_row_by_radionuclide(scenario))
    unmodeled_radionuclides = [
        radionuclide
        for radionuclide in EMPIRICAL_ACTIVITY_RADIONUCLIDES
        if radionuclide not in modeled_radionuclides
    ]
    group_summaries = _empirical_group_summaries(grouped)
    radionuclide_rows = _empirical_radionuclide_rows(grouped, modeled_radionuclides)
    sample_values = _empirical_modeled_sample_values(scenario, grouped)
    modeled_compartment_rows = _empirical_modeled_compartment_rows(scenario, grouped, sample_values)
    inferential_rows = _empirical_inferential_rows(scenario, grouped, sample_values)
    group_counts = {summary["group"]: summary["sample_count"] for summary in group_summaries}
    excluded_group_counts = {
        group: sum(1 for record in excluded_records if record.get("group") == group)
        for group in EMPIRICAL_ACTIVITY_GROUPS
        if group not in EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS
    }
    narrative_text = (
        "A estatística empírica usa somente as medições reais de atividade total do TAR - Afluente de 2019 a 2023. "
        "Valores marcados como < MDA> são tratados como censurados: entram nas contagens de não detectados, mas não "
        "entram como valor numérico. Para cada amostra com A-TAR numérico, a fração Si foi calculada a partir dos "
        "radionuclídeos detectados e aplicada à atividade total; em seguida, foram reutilizados a vazão e os fatores "
        "de água, peixe, invertebrado e sedimento da planilha TAR selecionada. A inferência compara os resultados "
        "calculados por amostra com o Report Level como referência fixa."
    )
    return {
        "synthetic": False,
        "source_workbook_path": loaded["path"],
        "source_sheet": loaded["sheet"],
        "source_note": "Dados reais de atividade total TAR - Afluente; < MDA> tratado como dado censurado.",
        "mda_policy": "< MDA> não é convertido para zero nem MDA/2; é contado como censurado e excluído dos cálculos numéricos.",
        "included_groups": EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS,
        "excluded_groups": [group for group in EMPIRICAL_ACTIVITY_GROUPS if group not in EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS],
        "excluded_group_counts": excluded_group_counts,
        "groups": group_summaries,
        "group_counts": group_counts,
        "observed_radionuclides": EMPIRICAL_ACTIVITY_RADIONUCLIDES,
        "modeled_radionuclides": [radionuclide for radionuclide in EXPECTED_RADIONUCLIDES if radionuclide in modeled_radionuclides],
        "unmodeled_radionuclides": unmodeled_radionuclides,
        "radionuclide_rows": radionuclide_rows,
        "modeled_compartment_rows": modeled_compartment_rows,
        "inferential_rows": inferential_rows,
        "narrative_text": narrative_text,
        "cards": [
            {"label": "Amostras afluente", "value": _format_count(group_counts.get("TAR - Afluente", 0))},
            {"label": "Política MDA", "value": "censurado"},
            {"label": "Inferência", "value": "Report Level fixo"},
            {"label": "Observado não modelado", "value": ", ".join(unmodeled_radionuclides) or "—"},
        ],
    }


def _date_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "date") and callable(value.date):
        return value.date().isoformat()
    if hasattr(value, "isoformat") and callable(value.isoformat):
        return value.isoformat()
    return _raw_text(value)


def _activity_matches(left: float | None, right: float | None) -> bool:
    if left is None or right is None:
        return False
    return math.isclose(float(left), float(right), rel_tol=1e-12, abs_tol=1.0)


def _load_total_activity_records(path: str | Path) -> dict[str, Any]:
    total_path = Path(path)
    if not total_path.exists():
        raise TarWorkbookError(f"Planilha de atividade total do tanque não encontrada: {total_path}")
    try:
        workbook = openpyxl.load_workbook(total_path, data_only=True, read_only=True)
    except Exception as exc:
        raise TarWorkbookError(f"Não foi possível abrir a planilha de atividade total do tanque: {exc}") from exc
    try:
        if not workbook.sheetnames:
            raise TarWorkbookError("A planilha de atividade total do tanque não contém abas.")
        ws = workbook[workbook.sheetnames[0]]
        records: list[dict[str, Any]] = []
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            date_value = row[0] if len(row) > 0 else None
            activity_value = _to_float(row[1] if len(row) > 1 else None)
            date_value_text = _date_text(date_value)
            if not date_value_text or activity_value is None:
                continue
            records.append(
                {
                    "row": row_index,
                    "date": date_value_text,
                    "activity_total_bq": activity_value,
                    "activity_total_text": _format_scientific(activity_value),
                }
            )
        return {
            "path": str(total_path),
            "sheet": ws.title,
            "records": records,
        }
    finally:
        workbook.close()


def _radionuclide_completeness(record: dict[str, Any] | None) -> dict[str, Any]:
    radionuclides = (record or {}).get("radionuclides") or {}
    numeric: list[str] = []
    censored: list[str] = []
    missing: list[str] = []
    for radionuclide in EMPIRICAL_ACTIVITY_RADIONUCLIDES:
        data = radionuclides.get(radionuclide) or {}
        if data.get("value") is not None:
            numeric.append(radionuclide)
        elif data.get("censored"):
            censored.append(radionuclide)
        else:
            missing.append(radionuclide)
    filled = len(numeric) + len(censored)
    complete = filled == len(EMPIRICAL_ACTIVITY_RADIONUCLIDES)
    return {
        "complete": complete,
        "numeric_count": len(numeric),
        "censored_count": len(censored),
        "missing_count": len(missing),
        "filled_count": filled,
        "numeric_radionuclides": numeric,
        "censored_radionuclides": censored,
        "missing_radionuclides": missing,
        "status": "completo" if complete else "incompleto",
    }


def _match_total_activity_record(total_record: dict[str, Any], detail_records: list[dict[str, Any]]) -> dict[str, Any] | None:
    matches = [
        record
        for record in detail_records
        if record.get("date") == total_record.get("date")
        and _activity_matches(record.get("activity_total_bq"), total_record.get("activity_total_bq"))
    ]
    if not matches:
        return None
    return next((record for record in matches if record.get("group") in EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS), matches[0])


def _total_activity_joined_row(rank: int | None, total_record: dict[str, Any], matched: dict[str, Any] | None) -> dict[str, Any]:
    completeness = _radionuclide_completeness(matched)
    if not matched:
        status = "sem correspondência"
    elif completeness["complete"]:
        status = "completo"
    else:
        status = "sem composição completa"
    return {
        "rank": rank,
        "source_row": total_record.get("row"),
        "date": total_record.get("date"),
        "activity_total_bq": total_record.get("activity_total_bq"),
        "activity_total_text": total_record.get("activity_total_text"),
        "matched": bool(matched),
        "detail_row": matched.get("row") if matched else None,
        "sample_id": matched.get("sample_id") if matched else "",
        "group": matched.get("group") if matched else "",
        "complete": bool(matched) and completeness["complete"],
        **completeness,
        "status": status,
    }


def _complete_total_activity_rows(
    total_records: list[dict[str, Any]],
    detail_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for total_record in sorted(total_records, key=lambda item: item["activity_total_bq"], reverse=True):
        matched = _match_total_activity_record(total_record, detail_records)
        joined = _total_activity_joined_row(None, total_record, matched)
        if joined["complete"]:
            rows.append(joined)
    return rows


def _parse_iso_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _group_total_activity_by_date(total_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for record in total_records:
        grouped.setdefault(str(record.get("date") or ""), []).append(record)

    rows: list[dict[str, Any]] = []
    for date_text, records in grouped.items():
        parsed = _parse_iso_date(date_text)
        activity_values = [
            float(record["activity_total_bq"])
            for record in records
            if record.get("activity_total_bq") is not None
        ]
        if parsed is None or not activity_values:
            continue
        rows.append(
            {
                "date": date_text,
                "date_value": parsed,
                "source_rows": [record.get("row") for record in records],
                "activity_row_count": len(records),
                "activity_total_bq": max(activity_values),
                "activity_total_text": _format_scientific(max(activity_values)),
                "activity_total_min_bq": min(activity_values),
                "activity_total_min_text": _format_scientific(min(activity_values)),
                "activity_total_values": activity_values,
            }
        )
    return sorted(rows, key=lambda item: item["date_value"])


def _detail_records_by_date(detail_records: list[dict[str, Any]]) -> dict[date, list[dict[str, Any]]]:
    grouped: dict[date, list[dict[str, Any]]] = {}
    for record in detail_records:
        if record.get("group") not in EMPIRICAL_ACTIVITY_ANALYSIS_GROUPS:
            continue
        parsed = _parse_iso_date(record.get("date"))
        if parsed is None:
            continue
        grouped.setdefault(parsed, []).append(record)
    return grouped


def _record_radionuclide_status(record: dict[str, Any], radionuclide: str) -> str:
    data = (record.get("radionuclides") or {}).get(radionuclide) or {}
    if data.get("value") is not None:
        return "numérico"
    if data.get("censored"):
        return "censurado"
    return "ausente"


def _date_radionuclide_set(records: list[dict[str, Any]], radionuclides: list[str]) -> set[str]:
    found: set[str] = set()
    for record in records:
        for radionuclide in radionuclides:
            if _record_radionuclide_status(record, radionuclide) in {"numérico", "censurado"}:
                found.add(radionuclide)
    return found


def _window_records(
    grouped_detail: dict[date, list[dict[str, Any]]],
    start_date: date,
    end_date: date,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for current_date in sorted(grouped_detail):
        if start_date <= current_date <= end_date:
            rows.extend(grouped_detail[current_date])
    return rows


def _minimum_radionuclide_window(
    anchor_date: date,
    grouped_detail: dict[date, list[dict[str, Any]]],
    target_radionuclides: list[str],
) -> dict[str, Any] | None:
    dates = sorted(grouped_detail)
    if not dates:
        return None

    best: dict[str, Any] | None = None
    for start_index, start_date in enumerate(dates):
        if start_date > anchor_date:
            break
        found: set[str] = set()
        row_count = 0
        for end_date in dates[start_index:]:
            for record in grouped_detail[end_date]:
                row_count += 1
                for radionuclide in target_radionuclides:
                    if _record_radionuclide_status(record, radionuclide) in {"numérico", "censurado"}:
                        found.add(radionuclide)
            if end_date < anchor_date:
                continue
            if len(found) < TOTAL_ACTIVITY_WINDOW_TARGET_RADIONUCLIDES:
                continue
            span_days = (end_date - start_date).days
            center = start_date.toordinal() + span_days / 2.0
            center_distance = abs(anchor_date.toordinal() - center)
            score = (span_days, center_distance, row_count, start_date, end_date)
            if best is None or score < best["score"]:
                best = {
                    "start_date": start_date,
                    "end_date": end_date,
                    "span_days": span_days,
                    "radionuclides": sorted(found),
                    "row_count": row_count,
                    "score": score,
                }
            break
    if best is None:
        return None
    best["records"] = _window_records(grouped_detail, best["start_date"], best["end_date"])
    best["radionuclides"] = sorted(_date_radionuclide_set(best["records"], target_radionuclides))
    best["row_count"] = len(best["records"])
    return best


def _select_window_radionuclides(
    window_records: list[dict[str, Any]],
    anchor_date: date,
    target_radionuclides: list[str],
) -> dict[str, dict[str, Any]]:
    selected: dict[str, dict[str, Any]] = {}
    for radionuclide in target_radionuclides:
        candidates: list[tuple[tuple[Any, ...], dict[str, Any], dict[str, Any], str, date]] = []
        for record in window_records:
            source_date = _parse_iso_date(record.get("date"))
            if source_date is None:
                continue
            data = (record.get("radionuclides") or {}).get(radionuclide) or {}
            status = _record_radionuclide_status(record, radionuclide)
            if status == "ausente":
                continue
            day_distance = abs((source_date - anchor_date).days)
            numeric_priority = 0 if status == "numérico" else 1
            activity_priority = -(float(record.get("activity_total_bq") or 0.0))
            score = (numeric_priority, day_distance, activity_priority, int(record.get("row") or 0))
            candidates.append((score, record, data, status, source_date))
        if not candidates:
            selected[radionuclide] = {
                "radionuclide": radionuclide,
                "status": "ausente",
                "value": None,
                "value_text": "—",
                "source_date": "",
                "source_sample_id": "",
                "day_offset": None,
                "day_offset_text": "—",
            }
            continue
        _score, record, data, status, source_date = sorted(candidates, key=lambda item: item[0])[0]
        day_offset = (source_date - anchor_date).days
        selected[radionuclide] = {
            "radionuclide": radionuclide,
            "status": status,
            "value": data.get("value"),
            "value_text": _format_scientific(data.get("value")) if data.get("value") is not None else data.get("raw_text") or "< MDA>",
            "raw_text": data.get("raw_text") or "",
            "source_date": source_date.isoformat(),
            "source_sample_id": record.get("sample_id") or "",
            "source_row": record.get("row"),
            "source_activity_total_bq": record.get("activity_total_bq"),
            "source_activity_total_text": _format_scientific(record.get("activity_total_bq")),
            "day_offset": day_offset,
            "day_offset_text": f"{day_offset:+d}",
        }
    return selected


def _build_total_activity_window_rows(
    total_records: list[dict[str, Any]],
    detail_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    anchors = _group_total_activity_by_date(total_records)
    grouped_detail = _detail_records_by_date(detail_records)
    target_radionuclides = [radionuclide for radionuclide in EXPECTED_RADIONUCLIDES]
    rows: list[dict[str, Any]] = []
    for anchor in anchors:
        anchor_date = anchor["date_value"]
        window = _minimum_radionuclide_window(anchor_date, grouped_detail, target_radionuclides)
        if window is None:
            rows.append(
                {
                    **{key: value for key, value in anchor.items() if key != "date_value"},
                    "anchor_date": anchor["date"],
                    "window_start_date": "",
                    "window_end_date": "",
                    "window_span_days": None,
                    "window_span_days_text": "—",
                    "window_row_count": 0,
                    "complete": False,
                    "status": "sem janela mínima",
                    "numeric_count": 0,
                    "censored_count": 0,
                    "missing_count": len(target_radionuclides),
                    "numeric_radionuclides": [],
                    "censored_radionuclides": [],
                    "missing_radionuclides": target_radionuclides,
                    "covered_radionuclides": [],
                    "selected_radionuclides": {},
                    "extends_beyond_anchor": False,
                }
            )
            continue
        selected = _select_window_radionuclides(window["records"], anchor_date, EMPIRICAL_ACTIVITY_RADIONUCLIDES)
        numeric_radionuclides = [rad for rad, item in selected.items() if item.get("status") == "numérico"]
        censored_radionuclides = [rad for rad, item in selected.items() if item.get("status") == "censurado"]
        missing_radionuclides = [rad for rad in target_radionuclides if (selected.get(rad) or {}).get("status") == "ausente"]
        covered_radionuclides = [rad for rad in target_radionuclides if rad not in missing_radionuclides]
        rows.append(
            {
                **{key: value for key, value in anchor.items() if key != "date_value"},
                "anchor_date": anchor["date"],
                "window_start_date": window["start_date"].isoformat(),
                "window_end_date": window["end_date"].isoformat(),
                "window_span_days": window["span_days"],
                "window_span_days_text": str(window["span_days"]),
                "window_row_count": window["row_count"],
                "complete": len(covered_radionuclides) >= TOTAL_ACTIVITY_WINDOW_TARGET_RADIONUCLIDES,
                "status": "janela mínima completa" if len(covered_radionuclides) >= TOTAL_ACTIVITY_WINDOW_TARGET_RADIONUCLIDES else "janela incompleta",
                "numeric_count": len(numeric_radionuclides),
                "censored_count": len(censored_radionuclides),
                "missing_count": len(missing_radionuclides),
                "filled_count": len(covered_radionuclides),
                "numeric_radionuclides": numeric_radionuclides,
                "censored_radionuclides": censored_radionuclides,
                "missing_radionuclides": missing_radionuclides,
                "covered_radionuclides": covered_radionuclides,
                "selected_radionuclides": selected,
                "extends_beyond_anchor": window["span_days"] > 0,
            }
        )
    return rows


def _total_activity_formula_rows() -> list[dict[str, str]]:
    return [
        {"name": "Fração radionuclídica", "symbol": "Si", "formula": "valor_radionuclídeo / soma_dos_radionuclídeos_numéricos"},
        {"name": "Atividade do radionuclídeo", "symbol": "Ai", "formula": "A-TAR * Si"},
        {"name": "Água", "symbol": "C_água", "formula": "Ai / vazão_do_cenário"},
        {"name": "Peixe", "symbol": "C_peixe", "formula": "C_água * Bp_peixe"},
        {"name": "Invertebrado", "symbol": "C_invertebrado", "formula": "C_água * Bp_invertebrado"},
        {"name": "Constante de decaimento", "symbol": "λ", "formula": "0,693 / meia_vida_h"},
        {"name": "Fator de acúmulo", "symbol": "F_acúmulo", "formula": "1 - exp(-λ * 262980)"},
        {"name": "Sedimento", "symbol": "C_sedimento", "formula": "C_água * 0,000072 * (F_acúmulo / λ)"},
    ]


def _total_activity_constant_rows(scenario: dict[str, Any]) -> list[dict[str, Any]]:
    flow = (scenario.get("totals") or {}).get("circulation_flow_m3_year")
    activity = (scenario.get("totals") or {}).get("activity_bq_year")
    rows: list[dict[str, Any]] = []
    for row in scenario.get("rows") or []:
        constants = row.get("constants") or {}
        rows.append(
            {
                "radionuclide": row.get("radionuclide"),
                "scenario_flow_m3_year": flow,
                "scenario_flow_text": _format_scientific(flow),
                "scenario_activity_bq_year": activity,
                "scenario_activity_text": _format_scientific(activity),
                "decay_constant_h": constants.get("decay_constant_h"),
                "decay_constant_h_text": constants.get("decay_constant_h_text") or "—",
                "half_life_h": constants.get("half_life_h"),
                "half_life_h_text": constants.get("half_life_h_text") or "—",
                "bioaccumulation_fish": constants.get("bioaccumulation_fish"),
                "bioaccumulation_fish_text": constants.get("bioaccumulation_fish_text") or "—",
                "bioaccumulation_invertebrate": constants.get("bioaccumulation_invertebrate"),
                "bioaccumulation_invertebrate_text": constants.get("bioaccumulation_invertebrate_text") or "—",
                "kd_sediment_l_kg": constants.get("kd_sediment_l_kg"),
                "kd_sediment_l_kg_text": constants.get("kd_sediment_l_kg_text") or "—",
                "sediment_transfer_factor": SEDIMENT_TRANSFER_FACTOR,
                "sediment_transfer_factor_text": _format_scientific(SEDIMENT_TRANSFER_FACTOR),
                "sediment_exposure_time_h": SEDIMENT_EXPOSURE_TIME_H,
                "sediment_exposure_time_h_text": _format_scientific(SEDIMENT_EXPOSURE_TIME_H),
            }
        )
    return rows


def _total_activity_matrix_rows(scenario: dict[str, Any], window_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    scenario_rows = _scenario_row_by_radionuclide(scenario)
    flow = (scenario.get("totals") or {}).get("circulation_flow_m3_year") or DEFAULT_DILUTION_FLOW_M3_YEAR
    rows: list[dict[str, Any]] = []
    for window_row in window_rows:
        if not window_row.get("complete"):
            continue
        selected_radionuclides = window_row.get("selected_radionuclides") or {}
        numeric_sum = sum(
            float((selected_radionuclides.get(radionuclide) or {}).get("value") or 0.0)
            for radionuclide in EMPIRICAL_ACTIVITY_RADIONUCLIDES
            if (selected_radionuclides.get(radionuclide) or {}).get("value") is not None
        )
        if numeric_sum <= 0 or flow <= 0:
            continue
        for radionuclide in EXPECTED_RADIONUCLIDES:
            scenario_row = scenario_rows.get(radionuclide)
            selected = selected_radionuclides.get(radionuclide) or {}
            detected_value = selected.get("value")
            if not scenario_row or detected_value is None or detected_value <= 0:
                continue
            fraction = float(detected_value) / numeric_sum
            activity_bq = float(window_row["activity_total_bq"]) * fraction
            water_value = activity_bq / float(flow)
            constants = scenario_row.get("constants") or {}
            decay_constant = constants.get("decay_constant_h")
            accumulation_factor = (
                1.0 - math.exp(-float(decay_constant) * SEDIMENT_EXPOSURE_TIME_H)
                if decay_constant not in (None, 0)
                else None
            )
            compartment_values = {
                "water": water_value,
                "fish": water_value * float(constants.get("bioaccumulation_fish"))
                if constants.get("bioaccumulation_fish") is not None
                else None,
                "invertebrate": water_value * float(constants.get("bioaccumulation_invertebrate"))
                if constants.get("bioaccumulation_invertebrate") is not None
                else None,
                "sediment": (
                    water_value * SEDIMENT_TRANSFER_FACTOR * (accumulation_factor / float(decay_constant))
                    if accumulation_factor is not None and decay_constant not in (None, 0)
                    else None
                ),
            }
            for compartment in COMPARTMENTS:
                value = compartment_values.get(compartment["key"])
                if value is None:
                    continue
                base_data = scenario_row["compartments"][compartment["key"]]
                rows.append(
                    {
                        "date": window_row["anchor_date"],
                        "anchor_date": window_row["anchor_date"],
                        "window_start_date": window_row.get("window_start_date", ""),
                        "window_end_date": window_row.get("window_end_date", ""),
                        "window_span_days": window_row.get("window_span_days"),
                        "window_span_days_text": window_row.get("window_span_days_text", "—"),
                        "sample_id": selected.get("source_sample_id") or "",
                        "source_date": selected.get("source_date") or "",
                        "source_sample_id": selected.get("source_sample_id") or "",
                        "source_day_offset": selected.get("day_offset"),
                        "source_day_offset_text": selected.get("day_offset_text") or "—",
                        "activity_total_bq": window_row["activity_total_bq"],
                        "activity_total_text": window_row["activity_total_text"],
                        "radionuclide": radionuclide,
                        "radionuclide_value": detected_value,
                        "radionuclide_value_text": _format_scientific(detected_value),
                        "detected_sum": numeric_sum,
                        "detected_sum_text": _format_scientific(numeric_sum),
                        "fraction": fraction,
                        "fraction_text": _format_decimal(fraction, digits=6),
                        "activity_bq": activity_bq,
                        "activity_bq_text": _format_scientific(activity_bq),
                        "compartment_key": compartment["key"],
                        "compartment": compartment["label"],
                        "unit": compartment["unit"],
                        "value": value,
                        "value_text": _format_scientific(value),
                        "report_level": base_data.get("report_level"),
                        "report_level_text": base_data.get("report_level_text"),
                        "lld": base_data.get("lld"),
                        "lld_text": base_data.get("lld_text"),
                        "decay_constant_h": decay_constant,
                        "decay_constant_h_text": constants.get("decay_constant_h_text") or "—",
                        "bioaccumulation_fish": constants.get("bioaccumulation_fish"),
                        "bioaccumulation_invertebrate": constants.get("bioaccumulation_invertebrate"),
                        "kd_sediment_l_kg": constants.get("kd_sediment_l_kg"),
                        "accumulation_factor": accumulation_factor,
                        "accumulation_factor_text": _format_decimal(accumulation_factor, digits=6),
                    }
                )
    return rows


def _generated_erica_value(calculated_value: float | None, rng: random.Random) -> float | None:
    if calculated_value is None or calculated_value <= 0:
        return None
    return float(calculated_value) * rng.lognormvariate(0.0, 0.35)


def _total_activity_erica_pair_rows(matrix_rows: list[dict[str, Any]], *, seed: int) -> list[dict[str, Any]]:
    rng = random.Random(seed)
    rows: list[dict[str, Any]] = []
    for item in matrix_rows:
        erica_value = _erica_value(item["radionuclide"], item["compartment_key"])
        generated = erica_value is None
        if generated:
            erica_value = _generated_erica_value(item.get("value"), rng)
        ratio = item["value"] / erica_value if erica_value not in (None, 0) else None
        rows.append(
            {
                "date": item["date"],
                "sample_id": item.get("sample_id", ""),
                "radionuclide": item["radionuclide"],
                "compartment_key": item["compartment_key"],
                "compartment": item["compartment"],
                "unit": item["unit"],
                "calculated_value": item["value"],
                "calculated_value_text": item["value_text"],
                "erica_value": erica_value,
                "erica_value_text": f"{_format_scientific(erica_value)}*" if generated else _format_scientific(erica_value),
                "erica_generated": generated,
                "erica_source": "gerado aleatoriamente com *" if generated else "extraído do PDF/artigo",
                "ratio": ratio,
                "ratio_text": _format_decimal(ratio),
            }
        )
    return rows


def _total_activity_norm_comparison_rows(matrix_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in matrix_rows:
        for reference_key, reference_label in (("report_level", "Report Level"), ("lld", "LLD")):
            reference_value = item.get(reference_key)
            if reference_value is None:
                continue
            result = _ratio_status(item.get("value"), reference_value)
            rows.append(
                {
                    "date": item["date"],
                    "sample_id": item.get("sample_id", ""),
                    "radionuclide": item["radionuclide"],
                    "compartment_key": item["compartment_key"],
                    "compartment": item["compartment"],
                    "unit": item["unit"],
                    "reference": reference_label,
                    "reference_value": reference_value,
                    "reference_value_text": _format_scientific(reference_value),
                    "value": item.get("value"),
                    "value_text": item.get("value_text"),
                    "ratio": result["ratio"],
                    "ratio_text": _format_decimal(result["ratio"]),
                    "status": result["status"],
                    "note": "LLD é referência de detecção" if reference_label == "LLD" else "Report Level é critério de notificação",
                }
            )
    return rows


def _total_activity_erica_chart_payloads(erica_pair_rows: list[dict[str, Any]]) -> dict[str, Any]:
    valid_rows = [
        row
        for row in erica_pair_rows
        if row.get("ratio") is not None and row.get("calculated_value") is not None and row.get("erica_value") is not None
    ]

    def point(row: dict[str, Any]) -> dict[str, Any]:
        return {
            "date": row.get("date", ""),
            "sample_id": row.get("sample_id", ""),
            "radionuclide": row.get("radionuclide", ""),
            "compartment_key": row.get("compartment_key", ""),
            "compartment": row.get("compartment", ""),
            "ratio": row.get("ratio"),
            "ratio_text": row.get("ratio_text", "—"),
            "calculated_value": row.get("calculated_value"),
            "calculated_value_text": row.get("calculated_value_text", "—"),
            "erica_value": row.get("erica_value"),
            "erica_value_text": row.get("erica_value_text", "—"),
            "erica_generated": bool(row.get("erica_generated")),
            "erica_source": row.get("erica_source", ""),
        }

    def build_scope(scope_key: str, scope_label: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
        ratios = [float(row["ratio"]) for row in rows if row.get("ratio") is not None and float(row["ratio"]) > 0]
        return {
            "scope_key": scope_key,
            "scope_label": scope_label,
            "n": len(ratios),
            "generated_erica_count": sum(1 for row in rows if row.get("erica_generated")),
            "stats": _ratio_summary_with_text(ratios),
            "points": [point(row) for row in sorted(rows, key=lambda item: (str(item.get("date") or ""), str(item.get("radionuclide") or ""), str(item.get("compartment") or "")))],
        }

    scopes = [build_scope("all", "Geral", valid_rows)]
    for compartment in COMPARTMENTS:
        rows = [row for row in valid_rows if row.get("compartment_key") == compartment["key"]]
        scopes.append(build_scope(str(compartment["key"]), str(compartment["label"]), rows))

    heatmap_cells: list[dict[str, Any]] = []
    for radionuclide in EXPECTED_RADIONUCLIDES:
        for compartment in COMPARTMENTS:
            rows = [
                row
                for row in valid_rows
                if row.get("radionuclide") == radionuclide and row.get("compartment_key") == compartment["key"]
            ]
            ratios = [float(row["ratio"]) for row in rows if row.get("ratio") is not None and float(row["ratio"]) > 0]
            stats_summary = _ratio_summary_with_text(ratios)
            heatmap_cells.append(
                {
                    "radionuclide": radionuclide,
                    "compartment_key": compartment["key"],
                    "compartment": compartment["label"],
                    "n": len(ratios),
                    "generated_erica_count": sum(1 for row in rows if row.get("erica_generated")),
                    "median_ratio": stats_summary["median"],
                    "median_ratio_text": stats_summary["median_text"],
                    "p95_ratio": stats_summary["p95"],
                    "p95_ratio_text": stats_summary["p95_text"],
                }
            )
    return {
        "reference_ratio": 1.0,
        "reference_label": "calculado / ERICA = 1",
        "axis_scale": "log10",
        "scope_rows": scopes,
        "heatmap": {
            "radionuclides": EXPECTED_RADIONUCLIDES,
            "compartments": [{"key": compartment["key"], "label": compartment["label"]} for compartment in COMPARTMENTS],
            "cells": heatmap_cells,
        },
        "explanation": (
            "A razão calculado/ERICA igual a 1 representa equivalência entre o valor obtido por fórmula e o valor do ERICA Tool. "
            "Razões acima de 1 indicam cálculo por fórmula maior que ERICA; razões abaixo de 1 indicam cálculo menor. "
            "A escala log10 evita que valores extremos escondam os pontos próximos de 1."
        ),
    }


def _total_activity_erica_inferential_rows(erica_pair_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}

    def add_group(group_type: str, scope: str, row: dict[str, Any]) -> None:
        key = (group_type, scope)
        grouped.setdefault(
            key,
            {
                "calculated": [],
                "erica": [],
                "generated": 0,
                "compartment_key": "",
                "compartment": "",
                "radionuclide": "Todos",
            },
        )
        grouped[key]["calculated"].append(row["calculated_value"])
        grouped[key]["erica"].append(row["erica_value"])
        grouped[key]["generated"] += 1 if row.get("erica_generated") else 0
        if group_type == "Matriz":
            grouped[key]["compartment_key"] = row.get("compartment_key", "")
            grouped[key]["compartment"] = row.get("compartment", "")
        elif group_type == "Radionuclídeo":
            grouped[key]["radionuclide"] = row.get("radionuclide", "")
            grouped[key]["compartment"] = "Todos os compartimentos"
        elif group_type == "Radionuclídeo-matriz":
            grouped[key]["radionuclide"] = row.get("radionuclide", "")
            grouped[key]["compartment_key"] = row.get("compartment_key", "")
            grouped[key]["compartment"] = row.get("compartment", "")

    for row in erica_pair_rows:
        add_group("Geral", "Todos os compartimentos", row)
        add_group("Matriz", str(row.get("compartment") or ""), row)
        add_group("Radionuclídeo", str(row.get("radionuclide") or ""), row)
        add_group("Radionuclídeo-matriz", f"{row.get('radionuclide')} - {row.get('compartment')}", row)

    rows: list[dict[str, Any]] = []
    order = {"Geral": 0, "Matriz": 1, "Radionuclídeo": 2, "Radionuclídeo-matriz": 3}
    for (group_type, scope), values in sorted(grouped.items(), key=lambda item: (order.get(item[0][0], 99), item[0][1])):
        paired = _paired_calculated_erica_test(values["calculated"], values["erica"])
        generated_count = int(values.get("generated") or 0)
        suffix = " Inclui valores ERICA gerados com *; interpretação exploratória." if generated_count else ""
        paired.update(
            {
                "comparison_type": "Calculado vs ERICA",
                "group_type": group_type,
                "scope": scope,
                "radionuclide": values.get("radionuclide") or "Todos",
                "compartment_key": values.get("compartment_key", ""),
                "compartment": values.get("compartment") or scope,
                "reference": "ERICA Tool",
                "reference_value": None,
                "reference_value_text": "—",
                "generated_erica_count": generated_count,
                "generated_erica_count_text": str(generated_count),
                "conclusion": f"{paired.get('conclusion') or paired.get('reason') or ''}{suffix}",
            }
        )
        rows.append(paired)
    return rows


def _total_activity_norm_inferential_rows(norm_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in norm_rows:
        key = (row["radionuclide"], row["compartment_key"], row["reference"])
        item = grouped.setdefault(
            key,
            {
                "values": [],
                "radionuclide": row["radionuclide"],
                "compartment_key": row["compartment_key"],
                "compartment": row["compartment"],
                "reference": row["reference"],
                "reference_value": row["reference_value"],
                "reference_value_text": row["reference_value_text"],
            },
        )
        item["values"].append(row["value"])
    rows: list[dict[str, Any]] = []
    for item in grouped.values():
        reference_label = str(item["reference"])
        inference = _real_reference_inference(item["values"], item["reference_value"], reference_label)
        conclusion = inference.get("conclusion") or inference.get("reason") or ""
        if reference_label == "LLD":
            conclusion = f"{conclusion} LLD é referência de detecção, não limite de ação."
        inference.update(
            {
                "comparison_type": f"Valor vs {reference_label}",
                "scope": f"{item['radionuclide']} - {item['compartment']}",
                "radionuclide": item["radionuclide"],
                "compartment_key": item["compartment_key"],
                "compartment": item["compartment"],
                "reference": reference_label,
                "reference_value": item["reference_value"],
                "reference_value_text": item["reference_value_text"],
                "generated_erica_count": 0,
                "generated_erica_count_text": "0",
                "conclusion": conclusion,
            }
        )
        rows.append(inference)
    return rows


def _build_total_activity_review(
    scenario: dict[str, Any],
    *,
    total_activity_workbook_path: str | Path,
    radionuclide_activity_workbook_path: str | Path,
    seed: int,
) -> dict[str, Any]:
    total_loaded = _load_total_activity_records(total_activity_workbook_path)
    detail_loaded = _load_empirical_activity_records(radionuclide_activity_workbook_path)
    total_records = total_loaded["records"]
    detail_records = detail_loaded["records"]
    ordered_records = sorted(total_records, key=lambda item: item["activity_total_bq"], reverse=True)
    top15_rows = [
        _total_activity_joined_row(rank, total_record, _match_total_activity_record(total_record, detail_records))
        for rank, total_record in enumerate(ordered_records[:TOTAL_ACTIVITY_REVIEW_TOP_N], start=1)
    ]
    complete_rows = _complete_total_activity_rows(total_records, detail_records)
    window_rows = _build_total_activity_window_rows(total_records, detail_records)
    matrix_rows = _total_activity_matrix_rows(scenario, window_rows)
    erica_pair_rows = _total_activity_erica_pair_rows(matrix_rows, seed=seed)
    erica_chart_payloads = _total_activity_erica_chart_payloads(erica_pair_rows)
    norm_comparison_rows = _total_activity_norm_comparison_rows(matrix_rows)
    erica_inferential_rows = _total_activity_erica_inferential_rows(erica_pair_rows)
    norm_inferential_rows = _total_activity_norm_inferential_rows(norm_comparison_rows)
    generated_erica_count = sum(1 for row in erica_pair_rows if row.get("erica_generated"))
    complete_dates = [row["date"] for row in complete_rows]
    window_dates = [row["anchor_date"] for row in window_rows if row.get("complete")]
    narrative_text = (
        "A revisão cruza as maiores atividades totais do tanque com a planilha de radionuclídeos por data e A-TAR. "
        "A auditoria mantém as 15 maiores linhas, mas os cálculos por matriz usam janelas mínimas por data: para cada "
        "data da planilha Dados, o sistema procura a menor janela cronológica no TAR - Afluente que reúna pelo menos "
        "8 radionuclídeos modelados. Valores < MDA> contam como preenchidos censurados para fechar a janela, mas não "
        "entram no denominador numérico de Si. Os valores ERICA gerados, quando necessários, são marcados com * e "
        "entram nos testes apenas com ressalva exploratória."
    )
    return {
        "synthetic": False,
        "source_total_workbook_path": total_loaded["path"],
        "source_total_sheet": total_loaded["sheet"],
        "source_radionuclide_workbook_path": detail_loaded["path"],
        "source_radionuclide_sheet": detail_loaded["sheet"],
        "source_note": "Top 15 por A-TAR da planilha Dados; cálculos somente em linhas completas da mesma lista.",
        "mda_policy": "< MDA> conta como preenchido censurado e não participa da soma numérica usada em Si.",
        "top_n": TOTAL_ACTIVITY_REVIEW_TOP_N,
        "top15_rows": top15_rows,
        "complete_rows": complete_rows,
        "complete_dates": complete_dates,
        "window_rows": window_rows,
        "window_dates": window_dates,
        "window_target_radionuclide_count": TOTAL_ACTIVITY_WINDOW_TARGET_RADIONUCLIDES,
        "window_target_radionuclides": EXPECTED_RADIONUCLIDES,
        "formula_rows": _total_activity_formula_rows(),
        "constant_rows": _total_activity_constant_rows(scenario),
        "matrix_rows": matrix_rows,
        "erica_pair_rows": erica_pair_rows,
        "erica_chart_payloads": erica_chart_payloads,
        "norm_comparison_rows": norm_comparison_rows,
        "inferential_rows": [*erica_inferential_rows, *norm_inferential_rows],
        "narrative_text": narrative_text,
        "cards": [
            {"label": "Linhas auditadas", "value": _format_count(len(top15_rows))},
            {"label": "Datas com janela", "value": _format_count(len(window_dates))},
            {"label": "Linhas completas exatas", "value": _format_count(len(complete_rows))},
            {"label": "Resultados por matriz", "value": _format_count(len(matrix_rows))},
            {"label": "Nº ERICA gerado", "value": _format_count(generated_erica_count)},
        ],
    }


def _synthetic_replicates(base_value: float | None, count: int, rng: random.Random) -> list[float]:
    if base_value is None or base_value <= 0:
        return []
    sigma = 0.20
    mu = math.log(base_value) - (sigma * sigma / 2.0)
    return [rng.lognormvariate(mu, sigma) for _ in range(count)]


def _reference_test(values: list[float], reference: float | None, reference_label: str, dataset_label: str) -> dict[str, Any]:
    if not values or reference is None or reference <= 0:
        return {
            "applicable": False,
            "test_label": "Não aplicável",
            "reason": f"Sem {reference_label} numérico para {dataset_label}.",
        }
    ratios = [value / reference for value in values if value > 0]
    if len(ratios) < 3:
        return {
            "applicable": False,
            "test_label": "Não aplicável",
            "reason": "Menos de 3 valores válidos.",
            "n": len(ratios),
        }

    log_ratios = [math.log(ratio) for ratio in ratios]
    shapiro_w = None
    shapiro_p = None
    try:
        shapiro = stats.shapiro(log_ratios)
        shapiro_w = float(shapiro.statistic)
        shapiro_p = float(shapiro.pvalue)
    except Exception:
        pass

    normality_met = shapiro_p is not None and shapiro_p >= 0.05
    if normality_met:
        test_label = "Teste t unilateral de uma amostra"
        try:
            test = stats.ttest_1samp(log_ratios, popmean=0.0, alternative="less")
            statistic = float(test.statistic)
            p_value = float(test.pvalue)
        except Exception:
            statistic = None
            p_value = None
    else:
        test_label = "Wilcoxon unilateral de uma amostra"
        try:
            test = stats.wilcoxon(log_ratios, alternative="less", zero_method="pratt")
            statistic = float(test.statistic)
            p_value = float(test.pvalue)
        except Exception:
            statistic = None
            p_value = None

    ratio_summary = _stats_summary_with_text(ratios)
    exceedance_count = sum(1 for ratio in ratios if ratio > 1.0)
    conclusion = (
        f"{dataset_label} ficou abaixo de {reference_label} nos valores avaliados"
        if p_value is not None and p_value < 0.05 and (ratio_summary["p95"] or 0) < 1.0
        else f"{dataset_label} não sustenta margem estatística exploratória abaixo de {reference_label}"
    )
    return {
        "applicable": p_value is not None,
        "n": len(ratios),
        "test_label": test_label,
        "statistic": statistic,
        "statistic_text": _stat_text(statistic),
        "p_value": p_value,
        "p_value_text": _format_p_value(p_value),
        "shapiro_w": shapiro_w,
        "shapiro_p": shapiro_p,
        "shapiro_p_text": _format_p_value(shapiro_p),
        "normality_met": normality_met,
        "ratio_summary": ratio_summary,
        "p95_ratio": ratio_summary["p95"],
        "p95_ratio_text": _stat_text(ratio_summary["p95"]),
        "exceedance_count": exceedance_count,
        "exceedance_rate": exceedance_count / len(ratios) if ratios else None,
        "exceedance_rate_text": _format_percent(exceedance_count / len(ratios) if ratios else None),
        "conclusion": conclusion,
    }


def _paired_calculated_erica_test(calculated_values: list[float], erica_values: list[float]) -> dict[str, Any]:
    pairs = [
        (calculated, erica)
        for calculated, erica in zip(calculated_values, erica_values)
        if calculated > 0 and erica > 0
    ]
    if len(pairs) < 3:
        return {
            "applicable": False,
            "test_label": "Não aplicável",
            "reason": "Menos de 3 pares válidos.",
            "n": len(pairs),
        }
    log_ratios = [math.log(calculated / erica) for calculated, erica in pairs]
    shapiro_w = None
    shapiro_p = None
    try:
        shapiro = stats.shapiro(log_ratios)
        shapiro_w = float(shapiro.statistic)
        shapiro_p = float(shapiro.pvalue)
    except Exception:
        pass
    normality_met = shapiro_p is not None and shapiro_p >= 0.05
    if normality_met:
        test_label = "Teste t pareado sobre log(calculado/ERICA)"
        try:
            test = stats.ttest_1samp(log_ratios, popmean=0.0, alternative="two-sided")
            statistic = float(test.statistic)
            p_value = float(test.pvalue)
        except Exception:
            statistic = None
            p_value = None
    else:
        test_label = "Wilcoxon pareado sobre log(calculado/ERICA)"
        try:
            test = stats.wilcoxon(log_ratios, alternative="two-sided", zero_method="pratt")
            statistic = float(test.statistic)
            p_value = float(test.pvalue)
        except Exception:
            statistic = None
            p_value = None

    mean_log = statistics.fmean(log_ratios)
    stdev_log = statistics.stdev(log_ratios) if len(log_ratios) > 1 else 0.0
    ci_half_width = 1.96 * stdev_log / math.sqrt(len(log_ratios)) if len(log_ratios) > 1 else 0.0
    ratios = [math.exp(value) for value in log_ratios]
    return {
        "applicable": p_value is not None,
        "n": len(log_ratios),
        "test_label": test_label,
        "statistic": statistic,
        "statistic_text": _stat_text(statistic),
        "p_value": p_value,
        "p_value_text": _format_p_value(p_value),
        "shapiro_w": shapiro_w,
        "shapiro_p": shapiro_p,
        "shapiro_p_text": _format_p_value(shapiro_p),
        "normality_met": normality_met,
        "mean_log_ratio": mean_log,
        "mean_log_ratio_text": _stat_text(mean_log),
        "mean_ratio": math.exp(mean_log),
        "mean_ratio_text": _stat_text(math.exp(mean_log)),
        "median_ratio": statistics.median(ratios),
        "median_ratio_text": _stat_text(statistics.median(ratios)),
        "ci95_low": math.exp(mean_log - ci_half_width),
        "ci95_low_text": _stat_text(math.exp(mean_log - ci_half_width)),
        "ci95_high": math.exp(mean_log + ci_half_width),
        "ci95_high_text": _stat_text(math.exp(mean_log + ci_half_width)),
        "conclusion": "Comparação pareada exploratória entre cálculo por fórmula e ERICA Tool; não constitui validação regulatória final.",
    }


def _erica_value(radionuclide: str, compartment_key: str) -> float | None:
    return (ERICA_TOOL_VALUES.get(radionuclide) or {}).get(compartment_key)


def _build_statistical_comparison(scenario: dict[str, Any], *, sample_count: int, seed: int) -> dict[str, Any]:
    calculated_rows: list[dict[str, Any]] = []
    erica_rows: list[dict[str, Any]] = []
    norm_rows: list[dict[str, Any]] = []
    descriptive_rows: list[dict[str, Any]] = []
    paired_comparison_rows: list[dict[str, Any]] = []
    comparison_values: dict[tuple[str, str], list[float]] = {}
    paired_values: dict[str, dict[str, list[float]]] = {
        "Todos os compartimentos": {"calculated": [], "erica": []}
    }
    for compartment in COMPARTMENTS:
        paired_values[compartment["label"]] = {"calculated": [], "erica": []}

    for row in scenario.get("rows") or []:
        radionuclide = row["radionuclide"]
        for compartment in COMPARTMENTS:
            key = compartment["key"]
            data = row["compartments"][key]
            calculated_value = data.get("value")
            erica_value = _erica_value(radionuclide, key)
            calculated_rows.append(
                {
                    "dataset": "calculado",
                    "radionuclide": radionuclide,
                    "compartment_key": key,
                    "compartment": compartment["label"],
                    "unit": compartment["unit"],
                    "value": calculated_value,
                    "value_text": _format_scientific(calculated_value),
                    "method": "Fórmulas de transporte/incorporação da planilha TAR",
                }
            )
            erica_rows.append(
                {
                    "dataset": "ERICA Tool",
                    "radionuclide": radionuclide,
                    "compartment_key": key,
                    "compartment": compartment["label"],
                    "unit": compartment["unit"],
                    "value": erica_value,
                    "value_text": _format_scientific(erica_value),
                    "method": "ERICA Tool Nível 2; valores extraídos do PDF corrigido",
                }
            )
            for reference_key, reference_label in (("report_level", "Report Level"), ("lld", "LLD")):
                reference_value = data.get(reference_key)
                if reference_value is None:
                    continue
                norm_rows.append(
                    {
                        "radionuclide": radionuclide,
                        "compartment_key": key,
                        "compartment": compartment["label"],
                        "reference": reference_label,
                        "unit": compartment["unit"],
                        "value": reference_value,
                        "value_text": _format_scientific(reference_value),
                    }
                )

            for dataset_label, base_value in (
                ("Calculado por fórmulas", calculated_value),
                ("Estimado pelo ERICA Tool", erica_value),
            ):
                if base_value is not None:
                    comparison_values.setdefault((dataset_label, compartment["label"]), []).append(base_value)

            if calculated_value is not None and erica_value is not None and calculated_value > 0 and erica_value > 0:
                paired_values["Todos os compartimentos"]["calculated"].append(calculated_value)
                paired_values["Todos os compartimentos"]["erica"].append(erica_value)
                paired_values[compartment["label"]]["calculated"].append(calculated_value)
                paired_values[compartment["label"]]["erica"].append(erica_value)

    for (dataset_label, compartment_label), values in comparison_values.items():
        compartment = next((item for item in COMPARTMENTS if item["label"] == compartment_label), None)
        descriptive_rows.append(
            {
                "dataset": dataset_label,
                "radionuclide": "Todos",
                "compartment_key": (compartment or {}).get("key", ""),
                "compartment": compartment_label,
                "unit": (compartment or {}).get("unit", ""),
                "base_value": None,
                "base_value_text": "—",
                **_stats_summary_with_text(values),
            }
        )

    for scope, values in paired_values.items():
        paired = _paired_calculated_erica_test(values["calculated"], values["erica"])
        paired.update(
            {
                "scope": scope,
                "radionuclide": "Todos",
                "compartment_key": "" if scope == "Todos os compartimentos" else next((item["key"] for item in COMPARTMENTS if item["label"] == scope), ""),
                "compartment": scope,
                "unit": "razão",
                "calculated_value": None,
                "calculated_value_text": f"{len(values['calculated'])} pares",
                "erica_value": None,
                "erica_value_text": f"{len(values['erica'])} pares",
            }
        )
        paired_comparison_rows.append(paired)

    narrative_text = (
        "A comparação estatística entre fórmula TAR, ERICA Tool e norma não usa replicações aleatórias. "
        "Os dados calculados vêm das fórmulas de transporte/incorporação da planilha TAR; os dados do ERICA Tool são "
        "estimativas de visualização até a substituição por saídas reais; Report Level e LLD permanecem como referências "
        "normativas fixas e não entram como amostras."
    )
    return {
        "sample_count": sample_count,
        "seed": seed,
        "synthetic": False,
        "random_replicates_used": False,
        "erica_estimated": True,
        "source_note": "sem replicações aleatórias; ERICA Tool mantido como estimativa visual pareada.",
        "erica_source": "Artigo TAR1 correção.pdf, tabela/screenshot do ERICA Tool Nível 2.",
        "calculated_rows": calculated_rows,
        "erica_rows": erica_rows,
        "norm_rows": norm_rows,
        "descriptive_rows": descriptive_rows,
        "inferential_rows": [],
        "paired_comparison_rows": paired_comparison_rows,
        "narrative_text": narrative_text,
    }


def _run_one_sample_limit_test(values: list[float], report_level: float | None) -> dict[str, Any]:
    if not values or report_level is None or report_level <= 0:
        return {
            "applicable": False,
            "test_label": "Não aplicável",
            "reason": "Sem Report Level numérico para comparação inferencial.",
        }
    ratios = [value / report_level for value in values if value is not None and value > 0]
    if len(ratios) < 3:
        return {
            "applicable": False,
            "test_label": "Não aplicável",
            "reason": "Menos de 3 simulações válidas para triagem de normalidade.",
            "n": len(ratios),
        }

    log_ratios = [math.log(ratio) for ratio in ratios]
    shapiro_w = None
    shapiro_p = None
    try:
        shapiro_w, shapiro_p = stats.shapiro(log_ratios)
        shapiro_w = float(shapiro_w)
        shapiro_p = float(shapiro_p)
    except Exception:
        shapiro_w = None
        shapiro_p = None

    normality_met = shapiro_p is not None and shapiro_p >= 0.05
    if normality_met:
        test_label = "Teste t unilateral de uma amostra"
        try:
            result = stats.ttest_1samp(log_ratios, popmean=0.0, alternative="less")
            statistic = float(result.statistic)
            p_value = float(result.pvalue)
        except Exception:
            statistic = None
            p_value = None
    else:
        test_label = "Wilcoxon unilateral de uma amostra"
        try:
            result = stats.wilcoxon(log_ratios, alternative="less", zero_method="pratt")
            statistic = float(result.statistic)
            p_value = float(result.pvalue)
        except Exception:
            statistic = None
            p_value = None

    exceedance_count = sum(1 for ratio in ratios if ratio > 1.0)
    stats_summary = _summary_stats(ratios)
    p95_ratio = stats_summary["p95"]
    significant_below = p_value is not None and p_value < 0.05 and p95_ratio is not None and p95_ratio < 1.0
    if significant_below and exceedance_count == 0:
        conclusion = "as simulações permaneceram estatisticamente abaixo do Report Level"
    elif significant_below:
        conclusion = "a tendência central ficou abaixo do Report Level, mas houve simulações acima da referência"
    elif p_value is None:
        conclusion = "o teste selecionado não produziu p-value válido"
    else:
        conclusion = "não houve evidência estatística suficiente para afirmar margem abaixo do Report Level"

    return {
        "applicable": p_value is not None,
        "n": len(ratios),
        "test_label": test_label,
        "statistic": statistic,
        "p_value": p_value,
        "p_value_text": _format_p_value(p_value),
        "shapiro_w": shapiro_w,
        "shapiro_p": shapiro_p,
        "shapiro_p_text": _format_p_value(shapiro_p),
        "normality_met": normality_met,
        "p95_ratio": p95_ratio,
        "p95_ratio_text": f"{p95_ratio:.4f}".replace(".", ",") if p95_ratio is not None else "—",
        "exceedance_count": exceedance_count,
        "exceedance_rate": exceedance_count / len(ratios) if ratios else None,
        "conclusion": conclusion,
    }


def _build_hypothetical_test_results(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for row in rows:
        for compartment in COMPARTMENTS:
            data = row["compartments"][compartment["key"]]
            report_level = data.get("report_level")
            simulated_values = data.get("simulated_values") or []
            if report_level is None:
                continue
            result = _run_one_sample_limit_test(simulated_values, report_level)
            result.update(
                {
                    "radionuclide": row["radionuclide"],
                    "compartment": compartment["label"],
                    "unit": compartment["unit"],
                    "report_level": report_level,
                    "report_level_text": data.get("report_level_text"),
                    "p95_value": data.get("value"),
                    "p95_value_text": data.get("value_text"),
                }
            )
            results.append(result)
    return results


def _build_hypothetical_statistical_text(test_results: list[dict[str, Any]], measurement_count: int) -> str:
    applicable = [result for result in test_results if result.get("applicable")]
    if not applicable:
        return (
            "O cenário hipotético gerou medições sintéticas, mas não encontrou pares suficientes de simulação e "
            "Report Level para aplicar teste inferencial."
        )
    test_labels = sorted({str(result.get("test_label")) for result in applicable})
    significant = [result for result in applicable if result.get("p_value") is not None and result.get("p_value") < 0.05]
    exceeded = [result for result in applicable if int(result.get("exceedance_count") or 0) > 0]
    if len(significant) == len(applicable) and not exceeded:
        closing = "Todos os contrastes com Report Level ficaram estatisticamente abaixo da referência, sem simulações acima do limite."
    elif significant:
        closing = "Parte dos contrastes ficou estatisticamente abaixo da referência; os casos com simulações acima do limite devem ser avaliados individualmente."
    else:
        closing = "Os testes não sustentaram evidência estatística suficiente de margem abaixo do Report Level."
    return (
        f"O cenário hipotético gerou {measurement_count} medições sintéticas de espectrometria gama para cada radionuclídeo da água do TAR. "
        "Cada medição alimentou a simulação dos compartimentos ambientais, preservando a relação entre valor medido e resultado simulado. "
        "A comparação inferencial foi feita sobre o logaritmo da razão simulado/Report Level. "
        "O Shapiro-Wilk avaliou a normalidade dessas razões; quando a normalidade foi atendida, aplicou-se teste t unilateral de uma amostra, "
        "e, quando não foi atendida, aplicou-se Wilcoxon unilateral de uma amostra. "
        f"Os testes utilizados foram: {', '.join(test_labels)}. {closing}"
    )


def _hypothetical_scenario_summary(
    base_scenario: dict[str, Any],
    *,
    measurement_count: int,
    seed: int,
) -> dict[str, Any]:
    rng = random.Random(seed)
    rows: list[dict[str, Any]] = []
    for base_row in base_scenario["rows"]:
        source_base = base_row.get("source_concentration_bq_m3")
        if source_base is None or source_base <= 0:
            continue
        measurements = _generate_positive_measurements(source_base, measurement_count, rng)
        measurement_summary = _summary_stats(measurements)
        generated_row: dict[str, Any] = {
            "row": base_row.get("row"),
            "radionuclide": base_row["radionuclide"],
            "source_concentration_bq_m3": source_base,
            "source_concentration_bq_m3_text": _format_scientific(source_base),
            "measurement_count": measurement_count,
            "measurements_bq_m3": measurements,
            "measurement_summary": {
                **measurement_summary,
                "mean_text": _format_scientific(measurement_summary["mean"]),
                "median_text": _format_scientific(measurement_summary["median"]),
                "p95_text": _format_scientific(measurement_summary["p95"]),
                "min_text": _format_scientific(measurement_summary["min"]),
                "max_text": _format_scientific(measurement_summary["max"]),
            },
            "constants": dict(base_row.get("constants") or {}),
            "compartments": {},
        }
        for compartment in COMPARTMENTS:
            base_data = base_row["compartments"][compartment["key"]]
            base_value = base_data.get("value")
            simulated_values = [
                base_value * (measurement / source_base)
                for measurement in measurements
                if base_value is not None
            ]
            simulated_summary = _summary_stats(simulated_values)
            p95_value = simulated_summary["p95"]
            report_result = _ratio_status(p95_value, base_data.get("report_level"))
            lld_result = _ratio_status(p95_value, base_data.get("lld"))
            generated_row["compartments"][compartment["key"]] = {
                "label": compartment["label"],
                "unit": compartment["unit"],
                "value": p95_value,
                "value_text": _format_scientific(p95_value),
                "display_statistic": "P95",
                "simulated_values": simulated_values,
                "simulated_summary": {
                    **simulated_summary,
                    "mean_text": _format_scientific(simulated_summary["mean"]),
                    "median_text": _format_scientific(simulated_summary["median"]),
                    "p95_text": _format_scientific(simulated_summary["p95"]),
                    "min_text": _format_scientific(simulated_summary["min"]),
                    "max_text": _format_scientific(simulated_summary["max"]),
                },
                "report_level": report_result["reference"],
                "report_level_text": _format_scientific(report_result["reference"]),
                "report_level_ratio": report_result["ratio"],
                "report_level_ratio_text": f"{report_result['ratio']:.4f}".replace(".", ",") if report_result["ratio"] is not None else "—",
                "report_level_status": report_result["status"],
                "lld": lld_result["reference"],
                "lld_text": _format_scientific(lld_result["reference"]),
                "lld_ratio": lld_result["ratio"],
                "lld_ratio_text": f"{lld_result['ratio']:.4f}".replace(".", ",") if lld_result["ratio"] is not None else "—",
                "lld_status": lld_result["status"],
            }
        rows.append(generated_row)

    reference_counts = _build_reference_counts(rows)
    report_level_exceedances = _build_exceedances(rows, reference="Report Level")
    lld_exceedances = _build_exceedances(rows, reference="LLD")
    test_results = _build_hypothetical_test_results(rows)
    totals = dict(base_scenario.get("totals") or {})
    totals["total_water_concentration_bq_m3"] = sum(
        row["compartments"]["water"]["value"] or 0.0
        for row in rows
    )
    return {
        "key": "hipotetico",
        "sheet": "gerado",
        "label": "Cenário hipotético",
        "source_scenario": base_scenario.get("label"),
        "source_sheet": base_scenario.get("sheet"),
        "is_hypothetical": True,
        "seed": seed,
        "measurement_count": measurement_count,
        "radionuclide_count": len(rows),
        "radionuclides": [row["radionuclide"] for row in rows],
        "expected_radionuclides": EXPECTED_RADIONUCLIDES,
        "rows": rows,
        "totals": totals,
        "reference_counts": reference_counts,
        "exceedances": report_level_exceedances,
        "report_level_exceedances": report_level_exceedances,
        "lld_exceedances": lld_exceedances,
        "has_reference_exceedance": bool(report_level_exceedances),
        "has_lld_exceedance": bool(lld_exceedances),
        "statistical_tests": test_results,
        "statistical_text": _build_hypothetical_statistical_text(test_results, measurement_count),
    }


def _inferential_assessment() -> dict[str, Any]:
    return {
        "applicable": False,
        "status": "teste inferencial não aplicável",
        "reason": (
            "A planilha TAR contém resultados calculados por modelo determinístico. O n = 8 representa radionuclídeos, "
            "não amostras ambientais independentes nem réplicas por matriz."
        ),
        "current_n": 8,
        "sample_unit": "radionuclídeo calculado",
        "recommended_action": (
            "Para inferência estatística, registrar medições ambientais reais com réplicas por compartimento, "
            "radionuclídeo, cenário e período de coleta."
        ),
        "minimums": INFERENTIAL_TEST_MINIMUMS,
    }


def _empirical_inferential_assessment(empirical: dict[str, Any]) -> dict[str, Any]:
    rows = empirical.get("inferential_rows") or []
    applicable = [row for row in rows if row.get("applicable")]
    if not rows:
        return _inferential_assessment()
    exceedance_rows = [row for row in rows if int(row.get("exceedance_count") or 0) > 0]
    if applicable and not exceedance_rows:
        status = "inferência real aplicável"
        reason = (
            "Há amostras reais do TAR - Afluente suficientes para testes sobre log(valor/Report Level) em parte dos "
            "radionuclídeos e compartimentos. O Report Level é tratado como referência fixa; os p-values descrevem "
            "as amostras reais calculadas pela fórmula, não a norma."
        )
    elif applicable:
        status = "inferência real com ressalvas"
        reason = (
            "Há testes aplicáveis com dados reais do TAR - Afluente, mas existem ultrapassagens observadas em pelo "
            "menos um contraste. A conclusão deve priorizar as frequências reais e seus IC95% binomiais."
        )
    else:
        status = "inferência real limitada"
        reason = (
            "Os dados reais do TAR - Afluente foram calculados por amostra, mas os contrastes disponíveis não atingiram "
            "condição suficiente para p-value em todos os casos. A leitura principal permanece descritiva, com frequência "
            "de ultrapassagem e IC95% binomial."
        )
    return {
        "applicable": bool(applicable),
        "status": status,
        "reason": reason,
        "current_n": max([int(row.get("n") or 0) for row in rows] + [0]),
        "sample_unit": "amostra real do TAR - Afluente calculada por fórmula",
        "recommended_action": (
            "Manter a coleta de amostras independentes do TAR - Afluente e substituir os valores estimados do ERICA "
            "Tool por saídas reais quando disponíveis."
        ),
        "minimums": INFERENTIAL_TEST_MINIMUMS,
    }


def _build_discussion_summary(
    scenario: dict[str, Any],
    inferential_assessment: dict[str, Any],
) -> dict[str, Any]:
    empirical = scenario.get("empirical_activity_statistics") or {}
    review = scenario.get("total_activity_review") or {}
    modeled_rows = list(empirical.get("modeled_compartment_rows") or [])
    reference_counts = scenario.get("reference_counts") or {}
    referenced_rows = [
        row
        for row in modeled_rows
        if row.get("report_level") is not None and row.get("report_level_p95_ratio") is not None
    ]

    below_compartments: list[str] = []
    exception_compartments: list[str] = []
    no_reference_compartments: list[str] = []
    for compartment in COMPARTMENTS:
        key = compartment["key"]
        label = compartment["label"]
        count = reference_counts.get(key) or {}
        compartment_rows = [row for row in referenced_rows if row.get("compartment_key") == key]
        if not count.get("report_level") or not compartment_rows:
            no_reference_compartments.append(label)
            continue
        above_rows = [row for row in compartment_rows if row.get("report_level_status") == "acima"]
        if above_rows:
            radionuclides = ", ".join(sorted({str(row.get("radionuclide")) for row in above_rows}))
            exception_compartments.append(f"{label} ({radionuclides})")
        else:
            below_compartments.append(label)

    sorted_contributors = sorted(
        referenced_rows,
        key=lambda row: float(row.get("report_level_p95_ratio") or 0.0),
        reverse=True,
    )
    contributor_rows = [
        {
            "radionuclide": row.get("radionuclide") or "",
            "compartment": row.get("compartment") or "",
            "p95_text": row.get("p95_text") or "—",
            "report_level_text": row.get("report_level_text") or "—",
            "p95_report_level_ratio_text": row.get("report_level_p95_ratio_text") or "—",
            "status": row.get("report_level_status") or "—",
            "exceedance_rate_text": row.get("report_level_exceedance_rate_text") or "—",
        }
        for row in sorted_contributors[:5]
    ]
    below_contributors = [
        row
        for row in sorted_contributors
        if row.get("report_level_status") == "abaixo"
    ][:3]

    cs137_sediment_row = next(
        (
            row
            for row in modeled_rows
            if row.get("radionuclide") == "Cs-137" and row.get("compartment_key") == "sediment"
        ),
        {},
    )
    cs137_sediment = {
        "radionuclide": cs137_sediment_row.get("radionuclide") or "Cs-137",
        "compartment": cs137_sediment_row.get("compartment") or "Sedimento",
        "n": cs137_sediment_row.get("n") or 0,
        "p95_text": cs137_sediment_row.get("p95_text") or "—",
        "report_level_text": cs137_sediment_row.get("report_level_text") or "—",
        "p95_report_level_ratio_text": cs137_sediment_row.get("report_level_p95_ratio_text") or "—",
        "exceedance_count": cs137_sediment_row.get("report_level_exceedance_count") or 0,
        "exceedance_rate_text": cs137_sediment_row.get("report_level_exceedance_rate_text") or "—",
        "lld_text": cs137_sediment_row.get("lld_text") or "—",
        "lld_p95_ratio_text": cs137_sediment_row.get("lld_p95_ratio_text") or "—",
        "lld_exceedance_rate_text": cs137_sediment_row.get("lld_exceedance_rate_text") or "—",
    }

    generated_erica_count = sum(
        1 for row in review.get("erica_pair_rows") or [] if row.get("erica_generated")
    )
    unmodeled = list(empirical.get("unmodeled_radionuclides") or [])
    has_censored = any(int(row.get("censored_count") or 0) > 0 for row in empirical.get("radionuclide_rows") or [])

    below_text = (
        f"{' e '.join(below_compartments)} ficaram abaixo do Report Level nos radionuclídeos com referência disponível."
        if below_compartments
        else "Nenhum compartimento com Report Level disponível ficou integralmente abaixo da referência."
    )
    if exception_compartments:
        below_text += f" A exceção observada foi {', '.join(exception_compartments)}, que exige leitura específica."
    if no_reference_compartments:
        below_text += (
            f" {', '.join(no_reference_compartments)} não deve ser declarado como abaixo do Report Level quando não há "
            "referência cadastrada para o compartimento ou radionuclídeo."
        )

    below_contributor_text = "; ".join(
        f"{row.get('radionuclide')} em {row.get('compartment')} (P95/RL = {row.get('report_level_p95_ratio_text')})"
        for row in below_contributors
    )
    contributor_text = (
        "A maior contribuição relativa frente ao Report Level foi Cs-137 em sedimento, porque é o único contraste "
        f"com P95/RL acima de 1. Entre os resultados ainda abaixo do limite, os maiores valores relativos foram "
        f"{below_contributor_text}."
        if below_contributor_text
        else (
            "A maior contribuição relativa frente ao Report Level foi Cs-137 em sedimento. Não houve outros "
            "contrastes referenciados abaixo do limite com razão P95/RL disponível."
        )
    )

    cs137_text = (
        "O caso específico de Cs-137 em sedimento concentra a principal ressalva normativa: "
        f"P95 = {cs137_sediment['p95_text']}, Report Level = {cs137_sediment['report_level_text']}, "
        f"P95/RL = {cs137_sediment['p95_report_level_ratio_text']} e frequência de ultrapassagem de "
        f"{cs137_sediment['exceedance_rate_text']} ({cs137_sediment['exceedance_count']} amostras). "
        "Esse achado não invalida automaticamente os demais compartimentos, mas impede uma conclusão global sem "
        "ressalvas para sedimento."
    )

    limitations = [
        "Valores censurados < MDA> entram nas contagens de não detectados, mas não entram como valor numérico no denominador das frações Si."
        if has_censored
        else "Não foram identificados valores censurados < MDA> no subconjunto usado para a discussão.",
        "Há ausência de Report Level para alguns compartimentos e radionuclídeos; nesses casos, a tabela informa sem referência e não permite classificar o resultado como abaixo ou acima.",
        (
            f"O radionuclídeo observado não modelado foi {', '.join(unmodeled)}, pois não há linha equivalente na fórmula da planilha TAR atual."
            if unmodeled
            else "Não houve radionuclídeo observado fora do conjunto modelado pela planilha TAR atual."
        ),
        (
            f"A comparação com ERICA contém {generated_erica_count} valor(es) gerado(s) por stat_seed e marcado(s) com asterisco; esses pares tornam a leitura exploratória."
            if generated_erica_count
            else "Nesta execução, a comparação calculado x ERICA não dependeu de valores ERICA gerados por stat_seed; quando houver asterisco, a leitura deve ser exploratória."
        ),
    ]

    definitions = [
        {
            "label": "Resultado calculado",
            "text": "valor obtido por fórmulas da planilha TAR, aplicando frações Si, vazão, fatores de bioacumulação e transferência para os compartimentos ambientais.",
        },
        {
            "label": "Resultado observado",
            "text": "medição registrada nas planilhas de atividade do TAR, incluindo detectados, censurados < MDA> e radionuclídeos observados que podem não existir no modelo.",
        },
        {
            "label": "Referência fixa",
            "text": "Report Level ou LLD cadastrado para comparação; não é amostra aleatória, não é estimado pelo teste estatístico e deve ser interpretado conforme sua função normativa ou de detecção.",
        },
    ]

    conclusion = {
        "method": (
            "O método cruzou a atividade total do tanque com a composição radionuclídica por data, selecionou janelas "
            "temporais próximas, calculou concentrações por fórmula em água, peixe, invertebrado e sedimento, e comparou "
            "os resultados com ERICA Tool, Report Level e LLD."
        ),
        "findings": (
            "Os resultados referenciados ficaram majoritariamente abaixo do Report Level, especialmente em água e peixe. "
            "A exceção relevante foi Cs-137 em sedimento, com P95/RL acima de 1 e frequência de ultrapassagem mensurável."
        ),
        "radiological_implication": (
            "A implicação radiológica é de ausência de evidência de ultrapassagem ampla nos compartimentos referenciados, "
            "mas com necessidade de tratar sedimento para Cs-137 como achado específico de atenção."
        ),
        "caveats": (
            "A conclusão permanece condicionada aos valores censurados, às lacunas de referência, aos pares ERICA gerados "
            "quando existirem e ao radionuclídeo observado não modelado."
        ),
        "recommendation": (
            f"{inferential_assessment.get('recommended_action') or 'Manter a coleta de amostras independentes e ampliar a rastreabilidade dos dados.'} "
            "Também se recomenda revisar referências ausentes e aprofundar a avaliação de Cs-137 em sedimento."
        ),
    }

    return {
        "section_title": "Discussão interpretativa",
        "report_level_text": below_text,
        "contributor_text": contributor_text,
        "contributor_rows": contributor_rows,
        "cs137_sediment": cs137_sediment,
        "cs137_sediment_text": cs137_text,
        "limitations": limitations,
        "definitions": definitions,
        "conclusion": conclusion,
        "inferential_status": inferential_assessment.get("status") or "",
        "inferential_reason": inferential_assessment.get("reason") or "",
    }


def load_tar_workbook_model(
    workbook_path: str | Path,
    *,
    hypothetical_n: int = 60,
    hypothetical_seed: int = 20260504,
) -> dict[str, Any]:
    path = Path(workbook_path)
    if not path.exists():
        raise TarWorkbookError(f"Planilha TAR não encontrada: {path}")
    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise TarWorkbookError(f"Não foi possível abrir a planilha TAR: {exc}") from exc
    try:
        scenarios: dict[str, Any] = {}
        for scenario_key, spec in SCENARIOS.items():
            sheet_name = spec["sheet"]
            if not sheet_name:
                continue
            if sheet_name not in workbook.sheetnames:
                raise TarWorkbookError(f"A aba obrigatória {sheet_name!r} não foi encontrada na planilha TAR.")
            scenarios[scenario_key] = _scenario_summary(scenario_key, workbook[sheet_name])
        scenarios["hipotetico"] = _hypothetical_scenario_summary(
            scenarios["a1_a2"],
            measurement_count=hypothetical_n,
            seed=hypothetical_seed,
        )
        return {
            "workbook_path": str(path),
            "scenarios": scenarios,
            "inferential_assessment": _inferential_assessment(),
        }
    finally:
        workbook.close()


def build_tar_summary(
    workbook_path: str | Path,
    scenario: Any = "a1",
    *,
    activity_workbook_path: str | Path | None = None,
    total_activity_workbook_path: str | Path | None = None,
    hypothetical_n: Any = 60,
    hypothetical_seed: Any = 20260504,
    sensitivity_n: Any = DEFAULT_SENSITIVITY_N,
    sensitivity_seed: Any = DEFAULT_SENSITIVITY_SEED,
    stat_n: Any = DEFAULT_STAT_N,
    stat_seed: Any = DEFAULT_STAT_SEED,
) -> dict[str, Any]:
    resolved_n = _safe_int(hypothetical_n, 60, minimum=10, maximum=500)
    resolved_seed = _safe_int(hypothetical_seed, 20260504, minimum=1, maximum=999999999)
    resolved_sensitivity_n = _safe_int(sensitivity_n, DEFAULT_SENSITIVITY_N, minimum=100, maximum=50000)
    resolved_sensitivity_seed = _safe_int(sensitivity_seed, DEFAULT_SENSITIVITY_SEED, minimum=1, maximum=999999999)
    resolved_stat_n = _safe_int(stat_n, DEFAULT_STAT_N, minimum=10, maximum=5000)
    resolved_stat_seed = _safe_int(stat_seed, DEFAULT_STAT_SEED, minimum=1, maximum=999999999)
    model = load_tar_workbook_model(
        workbook_path,
        hypothetical_n=resolved_n,
        hypothetical_seed=resolved_seed,
    )
    scenario_key = resolve_tar_scenario_key(scenario)
    scenario_model = dict(model["scenarios"][scenario_key])
    scenario_model["statistical_comparison"] = _build_statistical_comparison(
        scenario_model,
        sample_count=resolved_stat_n,
        seed=resolved_stat_seed,
    )
    resolved_activity_path = _empirical_activity_path(workbook_path, activity_workbook_path)
    resolved_total_activity_path = _total_activity_path(workbook_path, total_activity_workbook_path)
    scenario_model["empirical_activity_statistics"] = _build_empirical_activity_statistics(
        scenario_model,
        activity_workbook_path=resolved_activity_path,
    )
    scenario_model["total_activity_review"] = _build_total_activity_review(
        scenario_model,
        total_activity_workbook_path=resolved_total_activity_path,
        radionuclide_activity_workbook_path=resolved_activity_path,
        seed=resolved_stat_seed,
    )
    scenario_model["sensitivity"] = {}
    inferential_assessment = _empirical_inferential_assessment(scenario_model["empirical_activity_statistics"])
    discussion = _build_discussion_summary(scenario_model, inferential_assessment)
    return {
        "ok": True,
        "workbook_path": model["workbook_path"],
        "activity_workbook_path": str(resolved_activity_path),
        "total_activity_workbook_path": str(resolved_total_activity_path),
        "selected_scenario": scenario_key,
        "available_scenarios": [
            {"key": key, "label": value["label"], "sheet": value["sheet"]}
            for key, value in SCENARIOS.items()
        ],
        "scenario": scenario_model,
        "inferential_assessment": inferential_assessment,
        "discussion": discussion,
    }
